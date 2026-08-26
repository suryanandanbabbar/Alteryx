"""Tool Specifications Excel generator.

Produces an enterprise-grade, analyst-ready Excel workbook (.xlsx)
containing complete tool-by-tool specifications and data-flow explanations.
"""

from __future__ import annotations

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from awa.model.tool_specifications import ToolSpecificationsDocument

# Styling Constants
FONT_FAMILY = "Calibri"
COLOR_NAVY = "1B365D"       # Deep Navy Header
COLOR_WHITE = "FFFFFF"
COLOR_ZEBRA = "F8FAFC"      # Subtle Slate Tint
COLOR_BORDER = "CBD5E1"     # Light Slate Border
COLOR_TEXT_MAIN = "1E293B"  # Slate 800
COLOR_TEXT_BOLD = "0F172A"  # Slate 900

HEADER_FILL = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
ZEBRA_FILL = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")
WHITE_FILL = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")

HEADER_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_WHITE)
BODY_FONT = Font(name=FONT_FAMILY, size=9.5, bold=False, color=COLOR_TEXT_MAIN)
BOLD_BODY_FONT = Font(name=FONT_FAMILY, size=9.5, bold=True, color=COLOR_TEXT_BOLD)

THIN_BORDER_SIDE = Side(border_style="thin", color=COLOR_BORDER)
CELL_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
)

ALIGN_HEADER = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_WRAP_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER_TOP = Alignment(horizontal="center", vertical="top")


def generate_tool_specifications_excel(
    doc: ToolSpecificationsDocument,
    output_path: Path | str,
) -> None:
    """Generate a formatted Tool Specifications (.xlsx) workbook.

    Args:
        doc: Canonical ToolSpecificationsDocument containing tool rows.
        output_path: Path to save the resulting .xlsx workbook.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # PRIMARY WORKSHEET: Tool Specifications
    ws = wb.active
    ws.title = "Tool Specifications"
    ws.views.sheetView[0].showGridLines = True

    # 1. Header Row
    headers = [
        "Tool ID",
        "XML Tool Name",
        "Tool Type",
        "Role — What It Does",
        "Data Flow Explanation",
        "Input Tool",
        "Output Tool",
    ]

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_HEADER
        cell.border = CELL_BORDER

    ws.row_dimensions[1].height = 28

    # 2. Data Rows
    for row_idx, row_data in enumerate(doc.rows, start=2):
        is_zebra = (row_idx % 2 == 0)
        current_fill = ZEBRA_FILL if is_zebra else WHITE_FILL

        values = [
            row_data.tool_id_formatted,
            row_data.xml_tool_name,
            row_data.tool_type,
            row_data.role,
            row_data.data_flow_explanation,
            row_data.input_tool,
            row_data.output_tool,
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BOLD_BODY_FONT if col_idx == 1 else BODY_FONT
            cell.fill = current_fill
            cell.border = CELL_BORDER
            cell.alignment = ALIGN_WRAP_TOP

    # 3. Column Widths
    column_widths = {
        "A": 10,  # Tool ID
        "B": 42,  # XML Tool Name
        "C": 24,  # Tool Type
        "D": 55,  # Role — What It Does
        "E": 75,  # Data Flow Explanation
        "F": 28,  # Input Tool
        "G": 28,  # Output Tool
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 4. Freeze Header Row and Set Auto-Filter
    ws.freeze_panes = "A2"
    total_rows = max(len(doc.rows) + 1, 1)
    ws.auto_filter.ref = f"A1:G{total_rows}"

    wb.save(str(output_path))
