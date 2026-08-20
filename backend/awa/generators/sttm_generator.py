"""STTM (Source-to-Target Mapping) Excel generator.

Produces an enterprise-grade, analyst-ready Excel workbook (.xlsx)
containing complete field-level data lineage and transformation rules.
"""

from __future__ import annotations

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from awa.model.sttm import STTMDocument, STTMMapping

# Styling Constants
FONT_FAMILY = "Calibri"
COLOR_NAVY = "1B365D"       # Deep Navy Header
COLOR_WHITE = "FFFFFF"
COLOR_ZEBRA = "F8FAFC"      # Subtle Slate Tint
COLOR_BORDER = "CBD5E1"     # Light Slate Border
COLOR_MUTED = "64748B"      # Slate Gray text
COLOR_ACCENT = "0284C7"     # Accent Blue

HEADER_FILL = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
ZEBRA_FILL = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")
WHITE_FILL = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")

HEADER_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_WHITE)
BODY_FONT = Font(name=FONT_FAMILY, size=9.5, bold=False, color="1E293B")
BOLD_BODY_FONT = Font(name=FONT_FAMILY, size=9.5, bold=True, color="0F172A")
TITLE_FONT = Font(name=FONT_FAMILY, size=14, bold=True, color=COLOR_NAVY)
SUBTITLE_FONT = Font(name=FONT_FAMILY, size=10, italic=True, color=COLOR_MUTED)

THIN_BORDER_SIDE = Side(border_style="thin", color=COLOR_BORDER)
CELL_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
)

ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top")
ALIGN_CENTER_TOP = Alignment(horizontal="center", vertical="top")
ALIGN_WRAP_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_HEADER = Alignment(horizontal="left", vertical="center", wrap_text=True)


def generate_sttm_excel(
    sttm_doc: STTMDocument,
    output_path: Path,
) -> None:
    """Generate a formatted Source-to-Target Mapping (.xlsx) workbook.

    Args:
        sttm_doc: Canonical STTMDocument containing field-level mappings.
        output_path: Path to save the resulting .xlsx workbook.
    """
    wb = openpyxl.Workbook()

    # -------------------------------------------------------------
    # WORKSHEET 1: Source-to-Target Mapping
    # -------------------------------------------------------------
    ws_mappings = wb.active
    ws_mappings.title = "Source-to-Target Mapping"

    headers = [
        "Source Table",
        "Source Attribute",
        "Transformation",
        "Transformation Logic",
        "Target Table",
        "Target Attribute",
    ]

    # Write Header Row
    ws_mappings.row_dimensions[1].height = 28
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws_mappings.cell(row=1, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_HEADER
        cell.border = CELL_BORDER

    # Write Mapping Rows
    current_row = 2
    for idx, mapping in enumerate(sttm_doc.mappings):
        ws_mappings.row_dimensions[current_row].height = 36
        row_fill = ZEBRA_FILL if idx % 2 == 1 else WHITE_FILL

        row_values = [
            mapping.source_table,
            mapping.source_attribute,
            mapping.transformation,
            mapping.transformation_logic,
            mapping.target_table,
            mapping.target_attribute,
        ]

        for col_idx, val in enumerate(row_values, 1):
            cell = ws_mappings.cell(row=current_row, column=col_idx, value=val)
            cell.font = BOLD_BODY_FONT if col_idx in (1, 5, 6) else BODY_FONT
            cell.fill = row_fill
            cell.border = CELL_BORDER

            if col_idx == 4:  # Transformation Logic
                cell.alignment = ALIGN_WRAP_TOP
            elif col_idx == 3:  # Transformation Category
                cell.alignment = ALIGN_CENTER_TOP
            else:
                cell.alignment = ALIGN_LEFT_TOP

        current_row += 1

    # Freeze Header Row
    ws_mappings.freeze_panes = "A2"

    # Enable Autofilter
    max_row = max(current_row - 1, 1)
    ws_mappings.auto_filter.ref = f"A1:F{max_row}"

    # Set Column Widths
    col_widths = {
        "A": 26,  # Source Table
        "B": 26,  # Source Attribute
        "C": 22,  # Transformation
        "D": 58,  # Transformation Logic
        "E": 30,  # Target Table
        "F": 26,  # Target Attribute
    }
    for col_letter, width in col_widths.items():
        ws_mappings.column_dimensions[col_letter].width = width

    # Page Setup for Print / PDF Export
    ws_mappings.page_setup.orientation = ws_mappings.ORIENTATION_LANDSCAPE
    ws_mappings.sheet_properties.pageSetUpPr.fitToPage = True
    ws_mappings.page_setup.fitToWidth = 1
    ws_mappings.page_setup.fitToHeight = 0
    ws_mappings.print_title_rows = "1:1"

    # -------------------------------------------------------------
    # WORKSHEET 2: STTM Summary
    # -------------------------------------------------------------
    ws_summary = wb.create_sheet(title="STTM Summary")
    ws_summary.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_summary.cell(row=2, column=2, value="Source-to-Target Mapping Summary").font = TITLE_FONT
    ws_summary.cell(row=3, column=2, value=f"Workflow: {sttm_doc.workflow_name}").font = SUBTITLE_FONT

    # Metadata Table
    source_tables = sorted(list(set(m.source_table for m in sttm_doc.mappings)))
    target_tables = sorted(list(set(m.target_table for m in sttm_doc.mappings)))

    summary_metadata = [
        ("Workflow Name", sttm_doc.workflow_name),
        ("Total Field Mappings", len(sttm_doc.mappings)),
        ("Source Datasets / Tables", len(source_tables)),
        ("Target Datasets / Deliverables", len(target_tables)),
    ]

    ws_summary.cell(row=5, column=2, value="Metric").font = HEADER_FONT
    ws_summary.cell(row=5, column=2).fill = HEADER_FILL
    ws_summary.cell(row=5, column=2).border = CELL_BORDER

    ws_summary.cell(row=5, column=3, value="Value").font = HEADER_FONT
    ws_summary.cell(row=5, column=3).fill = HEADER_FILL
    ws_summary.cell(row=5, column=3).border = CELL_BORDER

    for i, (label, val) in enumerate(summary_metadata, start=6):
        c1 = ws_summary.cell(row=i, column=2, value=label)
        c1.font = BOLD_BODY_FONT
        c1.fill = WHITE_FILL if i % 2 == 0 else ZEBRA_FILL
        c1.border = CELL_BORDER

        c2 = ws_summary.cell(row=i, column=3, value=val)
        c2.font = BODY_FONT
        c2.fill = WHITE_FILL if i % 2 == 0 else ZEBRA_FILL
        c2.border = CELL_BORDER

    # Transformation Category Breakdown Table
    trans_counts: dict[str, int] = {}
    for m in sttm_doc.mappings:
        trans_counts[m.transformation] = trans_counts.get(m.transformation, 0) + 1

    start_breakdown_row = 12
    ws_summary.cell(row=start_breakdown_row, column=2, value="Transformation Category").font = HEADER_FONT
    ws_summary.cell(row=start_breakdown_row, column=2).fill = HEADER_FILL
    ws_summary.cell(row=start_breakdown_row, column=2).border = CELL_BORDER

    ws_summary.cell(row=start_breakdown_row, column=3, value="Mappings Count").font = HEADER_FONT
    ws_summary.cell(row=start_breakdown_row, column=3).fill = HEADER_FILL
    ws_summary.cell(row=start_breakdown_row, column=3).border = CELL_BORDER

    ws_summary.cell(row=start_breakdown_row, column=4, value="Share (%)").font = HEADER_FONT
    ws_summary.cell(row=start_breakdown_row, column=4).fill = HEADER_FILL
    ws_summary.cell(row=start_breakdown_row, column=4).border = CELL_BORDER

    total_m = len(sttm_doc.mappings)
    for j, (t_name, count) in enumerate(sorted(trans_counts.items(), key=lambda x: -x[1]), start=start_breakdown_row + 1):
        pct = f"{(count / total_m * 100):.1f}%" if total_m > 0 else "0.0%"
        row_f = WHITE_FILL if j % 2 == 0 else ZEBRA_FILL

        c1 = ws_summary.cell(row=j, column=2, value=t_name)
        c1.font = BODY_FONT
        c1.fill = row_f
        c1.border = CELL_BORDER

        c2 = ws_summary.cell(row=j, column=3, value=count)
        c2.font = BOLD_BODY_FONT
        c2.alignment = Alignment(horizontal="right")
        c2.fill = row_f
        c2.border = CELL_BORDER

        c3 = ws_summary.cell(row=j, column=4, value=pct)
        c3.font = BODY_FONT
        c3.alignment = Alignment(horizontal="right")
        c3.fill = row_f
        c3.border = CELL_BORDER

    # Column widths on summary tab
    ws_summary.column_dimensions["A"].width = 5
    ws_summary.column_dimensions["B"].width = 32
    ws_summary.column_dimensions["C"].width = 30
    ws_summary.column_dimensions["D"].width = 18

    # Ensure parent output directory exists and save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
