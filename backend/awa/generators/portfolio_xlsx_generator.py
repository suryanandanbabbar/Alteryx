"""ETL Portfolio Overview Excel workbook generator.

Produces an enterprise-grade, analyst-ready Excel workbook (.xlsx)
projecting canonical multi-workflow portfolio analysis, business features,
deterministic technical inventory, and ETL rationalisation intelligence.
"""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from awa.analysis.business_area_classifier import (
    ALLOWED_BUSINESS_AREAS,
    BUSINESS_AREA_DESCRIPTIONS,
)
from awa.model.analysis_result import CanonicalAnalysisResult
from awa.model.portfolio import (
    PortfolioAnalysis,
    PortfolioWorkflowSummary,
    RationalisationAnalysis,
)


# ---------------------------------------------------------------------------
# Reusable Styling Constants (Consistent with STTM & Tool Specs)
# ---------------------------------------------------------------------------

FONT_FAMILY = "Calibri"

COLOR_NAVY = "1B365D"        # Deep Navy Header
COLOR_NAVY_LIGHT = "2A4B7C"  # Secondary Navy
COLOR_WHITE = "FFFFFF"
COLOR_ZEBRA = "F8FAFC"       # Subtle Slate Tint (Slate-50)
COLOR_BORDER = "CBD5E1"      # Light Slate Border (Slate-300)
COLOR_TEXT_MAIN = "1E293B"   # Slate 800
COLOR_TEXT_BOLD = "0F172A"   # Slate 900
COLOR_MUTED = "64748B"       # Slate Gray
COLOR_CARD_BG = "F1F5F9"     # Slate 100
COLOR_ACCENT = "0284C7"      # Accent Blue (Sky 600)
COLOR_SUCCESS = "16A34A"     # Green 600
COLOR_WARNING = "D97706"     # Amber 600
COLOR_DANGER = "DC2626"      # Red 600

HEADER_FILL = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color=COLOR_NAVY_LIGHT, end_color=COLOR_NAVY_LIGHT, fill_type="solid")
ZEBRA_FILL = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")
WHITE_FILL = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")
CARD_FILL = PatternFill(start_color=COLOR_CARD_BG, end_color=COLOR_CARD_BG, fill_type="solid")

HEADER_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_WHITE)
SUBHEADER_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_WHITE)
BODY_FONT = Font(name=FONT_FAMILY, size=9.5, bold=False, color=COLOR_TEXT_MAIN)
BOLD_BODY_FONT = Font(name=FONT_FAMILY, size=9.5, bold=True, color=COLOR_TEXT_BOLD)
MUTED_FONT = Font(name=FONT_FAMILY, size=9, italic=True, color=COLOR_MUTED)

TITLE_FONT = Font(name=FONT_FAMILY, size=14, bold=True, color=COLOR_NAVY)
SUBTITLE_FONT = Font(name=FONT_FAMILY, size=9.5, italic=True, color=COLOR_MUTED)
KPI_NUM_FONT = Font(name=FONT_FAMILY, size=16, bold=True, color=COLOR_NAVY)
KPI_LABEL_FONT = Font(name=FONT_FAMILY, size=9, bold=True, color=COLOR_MUTED)

THIN_BORDER_SIDE = Side(border_style="thin", color=COLOR_BORDER)
CELL_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
)

ALIGN_HEADER = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_HEADER_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_WRAP_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER_TOP = Alignment(horizontal="center", vertical="top")
ALIGN_RIGHT_TOP = Alignment(horizontal="right", vertical="top")
ALIGN_KPI_NUM = Alignment(horizontal="center", vertical="center")
ALIGN_KPI_LABEL = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Helper Extraction & Derivation Functions
# ---------------------------------------------------------------------------

def _clean_str(val: Any, default: str = "") -> str:
    """Clean string or return default."""
    if val is None:
        return default
    s = str(val).strip()
    return s if s else default


def _format_lines(items: list[str]) -> str:
    """Format non-empty items as newline-separated text."""
    clean = [i.strip() for i in items if i and str(i).strip()]
    return "\n".join(clean) if clean else "Not determined from available evidence"


def _derive_business_outcome(w: PortfolioWorkflowSummary, res: CanonicalAnalysisResult | None) -> str:
    """Derive business outcome from why_it_matters and production targets."""
    if res and res.business_summary:
        if res.business_summary.why_it_matters and res.business_summary.why_it_matters.strip():
            return res.business_summary.why_it_matters.strip()
        # Fall back to target meanings
        meanings = [
            f"{out.name}: {out.business_meaning or out.likely_use}"
            for out in res.business_summary.business_outputs
            if out.business_meaning or out.likely_use
        ]
        if meanings:
            return "\n".join(meanings)
    if w.targets:
        return f"Generates {len(w.targets)} production deliverable(s): {', '.join(w.targets)}"
    return "Not determined from available evidence"


def _derive_key_features(res: CanonicalAnalysisResult | None) -> str:
    """Derive key features from stages and promoted rules."""
    if not res or not res.business_summary:
        return "Not determined from available evidence"

    features: list[str] = []
    # 1. Stages
    for stg in res.business_summary.processing_stages:
        summary_text = stg.summary or stg.name or stg.description
        if summary_text:
            features.append(f"{stg.stage_number}. {stg.name}: {summary_text}")

    # 2. Key promoted rules (up to 3)
    for rule in res.business_summary.business_rules[:3]:
        features.append(f"• {rule.rule_name} ({rule.category}): {rule.description}")

    return "\n".join(features) if features else "Not determined from available evidence"


def _normalize_tag(tag: str | None) -> str:
    """Normalize raw business-area tag to authoritative taxonomy bucket."""
    if not tag or str(tag).strip() in ("UNCLASSIFIED", ""):
        return "Other / Unclassified"
    return str(tag).strip()


def _derive_promoted_rules(res: CanonicalAnalysisResult | None) -> str:
    """Derive formatted business rules."""
    if not res or not res.business_summary or not res.business_summary.business_rules:
        return "No explicit business rules configured"

    lines: list[str] = []
    for r in res.business_summary.business_rules:
        lines.append(f"• {r.rule_name} [{r.category}]: {r.description}")
    return "\n".join(lines)


def _derive_input_domains(res: CanonicalAnalysisResult | None, w: PortfolioWorkflowSummary) -> str:
    """Derive formatted primary input domains."""
    if res and res.business_summary and res.business_summary.source_inputs:
        entries: list[str] = []
        for inp in res.business_summary.source_inputs:
            desc = f"{inp.name} ({inp.source_type})"
            if inp.business_role:
                desc += f" — {inp.business_role}"
            entries.append(desc)
        return "\n".join(entries)
    if w.sources:
        return "\n".join(w.sources)
    return "Not determined from available evidence"


def _derive_input_data_elements(res: CanonicalAnalysisResult | None, w: PortfolioWorkflowSummary) -> str:
    """Derive input tables/sheets and configured files."""
    elements: list[str] = []
    if res and res.business_summary and res.business_summary.source_inputs:
        for inp in res.business_summary.source_inputs:
            details: list[str] = [inp.source_filename or inp.name]
            if inp.sheet_or_table:
                details.append(f"Sheet/Table: {inp.sheet_or_table}")
            elements.append(" | ".join(details))
    elif w.sources:
        elements = list(w.sources)

    return "\n".join(elements) if elements else "Not determined from available evidence"


def _derive_primary_outputs(res: CanonicalAnalysisResult | None, w: PortfolioWorkflowSummary) -> str:
    """Derive production deliverables."""
    if res and res.business_summary and res.business_summary.business_outputs:
        outs = [f"{o.name} ({o.destination_type})" for o in res.business_summary.business_outputs]
        return "\n".join(outs)
    if w.targets:
        return "\n".join(w.targets)
    return "Not determined from available evidence"


def _derive_output_metrics(res: CanonicalAnalysisResult | None) -> str:
    """Derive output schema fields and metrics."""
    if not res:
        return "Not determined from available evidence"
    lines: list[str] = []
    if getattr(res, "sttm", None) and res.sttm and res.sttm.mappings:
        target_attrs = sorted({m.target_attribute for m in res.sttm.mappings if m.target_attribute})
        if target_attrs:
            display_attrs = target_attrs[:15]
            suffix = f" ... (+{len(target_attrs) - 15} more)" if len(target_attrs) > 15 else ""
            lines.append(f"Target Attributes: {', '.join(display_attrs)}{suffix}")
    if res.business_summary and res.business_summary.business_outputs:
        for out in res.business_summary.business_outputs:
            if out.sheet_or_table:
                lines.append(f"Destination: {out.name} ({out.sheet_or_table})")
    return "\n".join(lines) if lines else "Not determined from available evidence"


def _derive_business_consumers(w: PortfolioWorkflowSummary, res: CanonicalAnalysisResult | None) -> str:
    """Derive business consumers or explicit downstream workflow technical dependencies."""
    # Check if explicit business owner is documented
    if res and res.business_summary and res.business_summary.assessment:
        owner = res.business_summary.assessment.business_owner
        if owner and owner != "Not documented":
            return f"Business Owner / Consumer: {owner}"

    # Downstream workflows consuming this workflow's targets
    consumers = getattr(w, "downstream_consumers", [])
    if consumers:
        return f"Downstream workflows: {', '.join(consumers)}"

    return "Not determined from available evidence"


def _derive_customer_impact(w: PortfolioWorkflowSummary) -> str:
    """Derive customer impact only if supported by criticality factors or explicit evidence."""
    for factor in getattr(w, "criticality_factors", []):
        factor_lower = factor.lower()
        if "customer" in factor_lower or "policyholder" in factor_lower or "public" in factor_lower:
            return factor
    return "Not determined from available evidence"


def _derive_client_impact(w: PortfolioWorkflowSummary) -> str:
    """Derive internal stakeholder/client impact only if supported by criticality factors."""
    for factor in getattr(w, "criticality_factors", []):
        factor_lower = factor.lower()
        if "stakeholder" in factor_lower or "client" in factor_lower or "downstream" in factor_lower or "department" in factor_lower:
            return factor
    return "Not determined from available evidence"


def _derive_regulatory_relevance(res: CanonicalAnalysisResult | None) -> str:
    """Derive regulatory/compliance relevance only if explicit rules/stages mention it."""
    if not res or not res.business_summary:
        return "Not determined from available evidence"
    for r in res.business_summary.business_rules:
        r_text = f"{r.rule_name} {r.description}".lower()
        if any(term in r_text for term in ("statutory", "regulatory", "compliance", "audit", "solvency", "sox", "gdpr", "hipaa")):
            return f"Compliance rule: {r.rule_name} — {r.description}"
    for stg in res.business_summary.processing_stages:
        stg_text = f"{stg.name} {stg.summary} {stg.description}".lower()
        if any(term in stg_text for term in ("regulatory", "statutory", "compliance", "audit")):
            return f"Compliance stage: {stg.name} — {stg.summary}"
    return "Not determined from available evidence"


def _derive_reusability(w: PortfolioWorkflowSummary, rationalisation: RationalisationAnalysis | None) -> str:
    """Derive reusability potential from rationalisation candidates."""
    if rationalisation and rationalisation.candidates:
        for c in rationalisation.candidates:
            if w.workflow_id in c.workflow_ids or w.filename in c.workflow_names:
                other_names = [name for name in c.workflow_names if name != w.filename]
                others_str = ", ".join(other_names) if other_names else "other workflows"
                return f"High reusability: Shared operational logic identified with {others_str} ({c.recommendation_type})"
    return "Standalone workflow"


def _get_technical_facts(w: PortfolioWorkflowSummary, res: CanonicalAnalysisResult | None) -> dict[str, Any]:
    """Deterministically count and classify tools for technical inventory."""
    if not res or not res.workflow:
        return {
            "formula_count": 0,
            "join_count": 0,
            "filter_count": 0,
            "summarize_count": 0,
            "has_python": "No",
            "has_sql": "No",
            "has_macro": "No",
            "has_app": "No",
            "input_file_types": "None",
            "output_file_types": "None",
            "database_types": "None",
            "significant_tools": "Not available",
        }

    tools = res.workflow.tools.values()
    formula_count = sum(1 for t in tools if t.tool_type in ("Formula", "MultiRowFormula", "MultiFieldFormula"))
    join_count = sum(1 for t in tools if t.tool_type in ("Join", "JoinMultiple", "FindReplace", "FuzzyMatch", "AppendFields"))
    filter_count = sum(1 for t in tools if t.tool_type in ("Filter", "FilterInDB"))
    summarize_count = sum(1 for t in tools if t.tool_type in ("Summarize", "RunningTotal", "SummarizeInDB"))

    has_python = any("python" in (t.tool_type or "").lower() or "jupyter" in (t.tool_type or "").lower() for t in tools)
    has_macro = any("macro" in (t.tool_type or "").lower() or (t.plugin or "").lower().endswith(".yxmc") for t in tools)
    has_app = any(
        "question" in (t.tool_type or "").lower()
        or t.tool_type in ("Action", "Condition", "ControlParam", "DropDown", "ListBox", "RadioButton", "TextBox", "NumericUpDown")
        for t in tools
    )

    # SQL presence detection
    has_sql = False
    for t in tools:
        ttype_lower = (t.tool_type or "").lower()
        if "indb" in ttype_lower or "dynamicinput" in ttype_lower or "dynamicselect" in ttype_lower:
            has_sql = True
            break
        raw_xml = str(getattr(t.configuration, "raw_xml", "") or "").lower()
        if any(kw in raw_xml for kw in ("select ", "from ", "where ", "join ", "<table", "<query")):
            has_sql = True
            break

    # File types - safely parse without Path() on raw connection strings
    in_exts = set()
    for s in w.sources:
        try:
            clean_s = str(s).split("|||")[0].split("?")[0].strip()
            if "." in clean_s:
                ext = "." + clean_s.rsplit(".", 1)[-1].lower()
                if ext[1:].isalnum() and len(ext) <= 6:
                    in_exts.add(ext)
        except Exception:
            pass
    input_file_types = ", ".join(sorted(in_exts)) if in_exts else "None"

    out_exts = set()
    for t in w.targets:
        try:
            clean_t = str(t).split("|||")[0].split("?")[0].strip()
            if "." in clean_t:
                ext = "." + clean_t.rsplit(".", 1)[-1].lower()
                if ext[1:].isalnum() and len(ext) <= 6:
                    out_exts.add(ext)
        except Exception:
            pass
    output_file_types = ", ".join(sorted(out_exts)) if out_exts else "None"

    # Database types
    db_types = set()
    if has_sql:
        db_types.add("SQL Database")
    if any(ext in in_exts | out_exts for ext in (".xlsx", ".xls", ".xlsm")):
        db_types.add("Excel Workbook")
    if any(ext in in_exts | out_exts for ext in (".csv", ".tsv", ".txt")):
        db_types.add("Flat File (CSV)")
    if any(ext in in_exts | out_exts for ext in (".yxdb",)):
        db_types.add("Alteryx Database (.yxdb)")
    database_types = ", ".join(sorted(db_types)) if db_types else "File System"

    # Tool counts breakdown
    tool_type_counts: dict[str, int] = {}
    for t in tools:
        ttype = t.tool_type or "Unknown"
        tool_type_counts[ttype] = tool_type_counts.get(ttype, 0) + 1

    top_tools = sorted(tool_type_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    significant_tools = ", ".join(f"{tt} ({cnt})" for tt, cnt in top_tools)

    return {
        "formula_count": formula_count,
        "join_count": join_count,
        "filter_count": filter_count,
        "summarize_count": summarize_count,
        "has_python": "Yes" if has_python else "No",
        "has_sql": "Yes" if has_sql else "No",
        "has_macro": "Yes" if has_macro else "No",
        "has_app": "Yes" if has_app else "No",
        "input_file_types": input_file_types,
        "output_file_types": output_file_types,
        "database_types": database_types,
        "significant_tools": significant_tools,
    }


def _autofit_columns(ws: Any, max_override: dict[int, int] | None = None) -> None:
    """Compute responsive column widths with sensible bounds."""
    ws.views.sheetView[0].showGridLines = True
    max_override = max_override or {}

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            if "\n" in val_str:
                lines = val_str.split("\n")
                line_max = max(len(l) for l in lines) if lines else 0
                max_len = max(max_len, line_max)
            else:
                max_len = max(max_len, len(val_str))

        col_idx = col[0].column
        cap = max_override.get(col_idx, 45)
        calc_width = max(max_len + 3, 12)
        final_width = min(calc_width, cap)
        ws.column_dimensions[col_letter].width = final_width


# ---------------------------------------------------------------------------
# Sheet 1: Executive Summary
# ---------------------------------------------------------------------------

def _build_executive_summary_sheet(
    ws: Any,
    portfolio: PortfolioAnalysis,
    rationalisation: RationalisationAnalysis | None,
) -> None:
    """Build Sheet 1: Executive Summary with title banner, KPI cards, and full business area breakdown."""
    ws.title = "Executive Summary"

    # Title Banner
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value=f"ETL Portfolio Intelligence — {portfolio.portfolio_name}")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    # Subtitle / Scope
    ws.merge_cells("A2:F2")
    sub_text = (
        f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} | "
        f"Authoritative projection of canonical workflow IR, deterministic technical facts, and ETL rationalisation intelligence."
    )
    sub_cell = ws.cell(row=2, column=1, value=sub_text)
    sub_cell.font = SUBTITLE_FONT
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    # ---------------------------------------------------------
    # KPI Section (Row 4-6)
    # ---------------------------------------------------------
    high_criticality_count = sum(1 for w in portfolio.workflows if w.criticality_level == "HIGH")
    med_criticality_count = sum(1 for w in portfolio.workflows if w.criticality_level == "MEDIUM")
    low_criticality_count = sum(1 for w in portfolio.workflows if w.criticality_level == "LOW")

    high_complexity_count = sum(1 for w in portfolio.workflows if w.complexity_level == "HIGH")
    med_complexity_count = sum(1 for w in portfolio.workflows if w.complexity_level == "MEDIUM")
    low_complexity_count = sum(1 for w in portfolio.workflows if w.complexity_level == "LOW")

    total_opportunities = len(portfolio.rationalisation_candidates)
    if rationalisation and rationalisation.candidates:
        total_opportunities = len(rationalisation.candidates)

    kpis = [
        ("TOTAL WORKFLOWS", str(portfolio.metrics.total_workflows), COLOR_NAVY),
        ("SUCCESSFULLY ANALYSED", str(portfolio.metrics.successful_workflows), COLOR_SUCCESS),
        ("FAILED / PARTIAL", str(portfolio.metrics.failed_workflows), COLOR_DANGER if portfolio.metrics.failed_workflows > 0 else COLOR_MUTED),
        ("CRITICALITY PROFILE", f"H:{high_criticality_count} | M:{med_criticality_count} | L:{low_criticality_count}", COLOR_NAVY),
        ("COMPLEXITY PROFILE", f"H:{high_complexity_count} | M:{med_complexity_count} | L:{low_complexity_count}", COLOR_NAVY),
        ("RATIONALISATION CANDIDATES", str(total_opportunities), COLOR_ACCENT),
    ]

    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 28
    for idx, (label, val, color) in enumerate(kpis, 1):
        # KPI Label Row 4
        lbl_cell = ws.cell(row=4, column=idx, value=label)
        lbl_cell.font = KPI_LABEL_FONT
        lbl_cell.fill = CARD_FILL
        lbl_cell.alignment = ALIGN_KPI_LABEL
        lbl_cell.border = CELL_BORDER

        # KPI Value Row 5
        val_cell = ws.cell(row=5, column=idx, value=val)
        val_cell.font = Font(name=FONT_FAMILY, size=15, bold=True, color=color)
        val_cell.fill = CARD_FILL
        val_cell.alignment = ALIGN_KPI_NUM
        val_cell.border = CELL_BORDER

    # ---------------------------------------------------------
    # Business Area Breakdown Table (Row 8 onwards)
    # ---------------------------------------------------------
    table_start_row = 8
    ws.cell(row=table_start_row, column=1, value="BUSINESS AREA PORTFOLIO DISTRIBUTION").font = Font(name=FONT_FAMILY, size=11, bold=True, color=COLOR_NAVY)
    ws.row_dimensions[table_start_row].height = 24

    headers = [
        "Business Area",
        "Workflow Count",
        "Primary Function Focus",
        "Description",
    ]

    header_row = table_start_row + 1
    ws.row_dimensions[header_row].height = 26
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_HEADER_CENTER if col_idx == 2 else ALIGN_HEADER
        cell.border = CELL_BORDER

    # Collect all configured business areas + Other / Unclassified in canonical order
    all_areas = [
        "Claims & Risk",
        "Legal",
        "Underwriting",
        "Sales & Distribution",
        "Actuarial",
        "Other / Unclassified",
    ]

    current_row = header_row + 1
    for idx, area in enumerate(all_areas):
        # Workflows assigned to this area
        area_wfs = [
            w for w in portfolio.workflows
            if (w.business_area_tag == area or (area == "Other / Unclassified" and w.business_area_tag in ("UNCLASSIFIED", "", None)))
        ]
        cnt = len(area_wfs)

        # Derive primary function focus
        funcs = [w.business_function for w in area_wfs if w.business_function and w.business_function != "Unassigned"]
        primary_func = funcs[0] if funcs else "No workflows currently mapped"

        desc = BUSINESS_AREA_DESCRIPTIONS.get(area, "Unclassified or custom business area.")

        row_fill = ZEBRA_FILL if idx % 2 == 1 else WHITE_FILL
        ws.row_dimensions[current_row].height = 36

        # Col 1: Area Name
        c1 = ws.cell(row=current_row, column=1, value=area)
        c1.font = BOLD_BODY_FONT
        c1.fill = row_fill
        c1.alignment = ALIGN_WRAP_TOP
        c1.border = CELL_BORDER

        # Col 2: Count
        c2 = ws.cell(row=current_row, column=2, value=cnt)
        c2.font = BODY_FONT
        c2.fill = row_fill
        c2.alignment = ALIGN_CENTER_TOP
        c2.number_format = "#,##0"
        c2.border = CELL_BORDER

        # Col 3: Primary Function Focus
        c3 = ws.cell(row=current_row, column=3, value=primary_func)
        c3.font = BODY_FONT
        c3.fill = row_fill
        c3.alignment = ALIGN_WRAP_TOP
        c3.border = CELL_BORDER

        # Col 4: Description
        c4 = ws.cell(row=current_row, column=4, value=desc)
        c4.font = BODY_FONT
        c4.fill = row_fill
        c4.alignment = ALIGN_WRAP_TOP
        c4.border = CELL_BORDER

        current_row += 1

    # Freeze at table headers
    ws.freeze_panes = f"A{header_row + 1}"
    _autofit_columns(ws, {1: 24, 2: 16, 3: 38, 4: 65})


# ---------------------------------------------------------------------------
# Sheet 2: Portfolio Summary / Inventory
# ---------------------------------------------------------------------------

def _build_portfolio_summary_sheet(
    ws: Any,
    portfolio: PortfolioAnalysis,
    successful_results: dict[str, CanonicalAnalysisResult],
) -> None:
    """Build Sheet 2: Inventory with one row per analysed workflow."""
    ws.title = "Inventory"

    headers = [
        "Workflow Name",
        "Business Area",
        "Business Function",
        "Business Purpose",
        "Business Outcome",
        "Key Business Features",
        "Primary Inputs",
        "Primary Outputs",
        "Business Consumers",
        "Customer Impact",
        "Client Impact",
        "Business Scope",
        "Criticality",
        "Criticality Score",
        "Criticality Justification",
        "Complexity",
        "Tool Count",
        "Connection Count",
        "Input Count",
        "Output Count",
        "Last Run",
        "Frequency",
        "Business Area Tag Source",
    ]

    ws.row_dimensions[1].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_HEADER_CENTER if col_idx in (13, 14, 16, 17, 18, 19, 20, 21, 22, 23) else ALIGN_HEADER
        cell.border = CELL_BORDER

    current_row = 2
    for idx, w in enumerate(portfolio.workflows):
        res = successful_results.get(w.workflow_id)
        row_fill = ZEBRA_FILL if idx % 2 == 1 else WHITE_FILL
        ws.row_dimensions[current_row].height = 42

        # Safe values
        biz_purpose = w.business_purpose or (
            f"Workflow analysis failed: {w.error_message}" if w.status == "FAILED" else "Not determined from available evidence"
        )
        biz_outcome = _derive_business_outcome(w, res)
        key_features = _derive_key_features(res)
        primary_inputs = "\n".join(w.sources) if w.sources else "Not determined from available evidence"
        primary_outputs = "\n".join(w.targets) if w.targets else "Not determined from available evidence"
        biz_consumers = _derive_business_consumers(w, res)
        cust_impact = _derive_customer_impact(w)
        client_impact = _derive_client_impact(w)
        biz_scope = "Not determined from available evidence"

        # Authoritative Last Run from canonical metadata
        last_run = getattr(w, "last_run", None)
        if not last_run or last_run == "Not documented":
            if hasattr(w, "factor_assessments") and isinstance(w.factor_assessments, dict):
                last_run = w.factor_assessments.get("last_run", {}).get("display_value")
        if not last_run or last_run == "Not documented":
            if res and res.business_summary and res.business_summary.factor_assessments:
                last_run = res.business_summary.factor_assessments.get("last_run", {}).get("display_value")
        if not last_run or last_run == "Not documented":
            if res and res.workflow and res.workflow.metadata and res.workflow.metadata.properties:
                props = res.workflow.metadata.properties
                meta = props.get("MetaInfo", {}) if isinstance(props.get("MetaInfo"), dict) else {}
                last_run = (
                    props.get("last_run")
                    or props.get("last_executed")
                    or props.get("last_run_date")
                    or props.get("lastRun")
                    or props.get("LastRun")
                    or meta.get("last_run")
                    or meta.get("lastRun")
                    or meta.get("LastRun")
                )
        if not last_run or not str(last_run).strip():
            last_run = "Not documented"

        # Authoritative Frequency from canonical deterministic metadata
        freq = getattr(w, "frequency", None)
        if not freq or freq == "Not documented":
            if hasattr(w, "factor_assessments") and isinstance(w.factor_assessments, dict):
                freq = w.factor_assessments.get("frequency", {}).get("display_value")
        if not freq or freq == "Not documented":
            if res and res.business_summary and res.business_summary.factor_assessments:
                freq = res.business_summary.factor_assessments.get("frequency", {}).get("display_value")
        if not freq or freq == "Not documented":
            if res and res.workflow and res.workflow.metadata and res.workflow.metadata.properties:
                props = res.workflow.metadata.properties
                meta = props.get("MetaInfo", {}) if isinstance(props.get("MetaInfo"), dict) else {}
                freq = (
                    props.get("frequency")
                    or props.get("schedule")
                    or props.get("run_frequency")
                    or props.get("frequency_schedule")
                    or props.get("Frequency")
                    or meta.get("frequency")
                    or meta.get("Frequency")
                )
        if not freq or not str(freq).strip():
            freq = "Not documented"

        row_values = [
            (w.filename, BOLD_BODY_FONT, ALIGN_WRAP_TOP, None),
            (_normalize_tag(w.business_area_tag), BODY_FONT, ALIGN_WRAP_TOP, None),
            (w.business_function or "Unassigned", BODY_FONT, ALIGN_WRAP_TOP, None),
            (biz_purpose, BODY_FONT, ALIGN_WRAP_TOP, None),
            (biz_outcome, BODY_FONT, ALIGN_WRAP_TOP, None),
            (key_features, BODY_FONT, ALIGN_WRAP_TOP, None),
            (primary_inputs, BODY_FONT, ALIGN_WRAP_TOP, None),
            (primary_outputs, BODY_FONT, ALIGN_WRAP_TOP, None),
            (biz_consumers, BODY_FONT, ALIGN_WRAP_TOP, None),
            (cust_impact, BODY_FONT, ALIGN_WRAP_TOP, None),
            (client_impact, BODY_FONT, ALIGN_WRAP_TOP, None),
            (biz_scope, BODY_FONT, ALIGN_WRAP_TOP, None),
            (w.criticality_level, BOLD_BODY_FONT, ALIGN_CENTER_TOP, None),
            (round(w.criticality_score, 1), BODY_FONT, ALIGN_CENTER_TOP, "0.0"),
            (w.criticality_justification or "Not determined from available evidence", BODY_FONT, ALIGN_WRAP_TOP, None),
            (w.complexity_level, BOLD_BODY_FONT, ALIGN_CENTER_TOP, None),
            (w.node_count, BODY_FONT, ALIGN_CENTER_TOP, "#,##0"),
            (w.connection_count, BODY_FONT, ALIGN_CENTER_TOP, "#,##0"),
            (w.source_count, BODY_FONT, ALIGN_CENTER_TOP, "#,##0"),
            (w.target_count, BODY_FONT, ALIGN_CENTER_TOP, "#,##0"),
            (last_run, BODY_FONT, ALIGN_CENTER_TOP, None),
            (freq, BODY_FONT, ALIGN_CENTER_TOP, None),
            (w.business_area_tag_source, BODY_FONT, ALIGN_CENTER_TOP, None),
        ]

        for col_idx, (val, font, alignment, num_format) in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font
            cell.fill = row_fill
            cell.alignment = alignment
            cell.border = CELL_BORDER
            if num_format:
                cell.number_format = num_format

        current_row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autofit_columns(ws, {1: 28, 2: 20, 3: 24, 4: 45, 5: 45, 6: 45, 7: 30, 8: 30, 9: 30, 10: 30, 11: 30, 12: 24, 13: 15, 14: 16, 15: 45, 16: 15, 21: 18, 22: 18, 23: 24})


# ---------------------------------------------------------------------------
# Sheet 3: Technical Inventory
# ---------------------------------------------------------------------------

def _build_technical_inventory_sheet(
    ws: Any,
    portfolio: PortfolioAnalysis,
    successful_results: dict[str, CanonicalAnalysisResult],
) -> None:
    """Build Sheet 4: Technical Inventory populated deterministically from canonical tool taxonomy."""
    ws.title = "Technical Inventory"

    headers = [
        "Workflow Name",
        "Business Area",
        "Tool Count",
        "Connection Count",
        "Input Count",
        "Output Count",
        "Formula Count",
        "Join Count",
        "Filter Count",
        "Summarize / Aggregation Count",
        "Python Tool Present",
        "SQL Present",
        "Macro Present",
        "Workflow App / Interface Tools",
        "Input File Types",
        "Output File Types",
        "Database / Data Source Types",
        "Tool Types / Significant Tool Inventory",
    ]

    ws.row_dimensions[1].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_HEADER_CENTER if 3 <= col_idx <= 14 else ALIGN_HEADER
        cell.border = CELL_BORDER

    current_row = 2
    for idx, w in enumerate(portfolio.workflows):
        res = successful_results.get(w.workflow_id)
        facts = _get_technical_facts(w, res)
        row_fill = ZEBRA_FILL if idx % 2 == 1 else WHITE_FILL
        ws.row_dimensions[current_row].height = 36

        row_values = [
            (w.filename, BOLD_BODY_FONT, ALIGN_WRAP_TOP, None),
            (_normalize_tag(w.business_area_tag), BODY_FONT, ALIGN_WRAP_TOP, None),
            (w.node_count, BODY_FONT, ALIGN_CENTER_TOP, "#,##0"),
            (w.connection_count, BODY_FONT, ALIGN_CENTER_TOP, "#,##0"),
            (w.source_count, BODY_FONT, ALIGN_CENTER_TOP, "#,##0"),
            (w.target_count, BODY_FONT, ALIGN_CENTER_TOP, "#,##0"),
            (facts["formula_count"], BODY_FONT, ALIGN_CENTER_TOP, "#,##0"),
            (facts["join_count"], BODY_FONT, ALIGN_CENTER_TOP, "#,##0"),
            (facts["filter_count"], BODY_FONT, ALIGN_CENTER_TOP, "#,##0"),
            (facts["summarize_count"], BODY_FONT, ALIGN_CENTER_TOP, "#,##0"),
            (facts["has_python"], BOLD_BODY_FONT, ALIGN_CENTER_TOP, None),
            (facts["has_sql"], BOLD_BODY_FONT, ALIGN_CENTER_TOP, None),
            (facts["has_macro"], BOLD_BODY_FONT, ALIGN_CENTER_TOP, None),
            (facts["has_app"], BOLD_BODY_FONT, ALIGN_CENTER_TOP, None),
            (facts["input_file_types"], BODY_FONT, ALIGN_WRAP_TOP, None),
            (facts["output_file_types"], BODY_FONT, ALIGN_WRAP_TOP, None),
            (facts["database_types"], BODY_FONT, ALIGN_WRAP_TOP, None),
            (facts["significant_tools"], BODY_FONT, ALIGN_WRAP_TOP, None),
        ]

        for col_idx, (val, font, alignment, num_format) in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font
            cell.fill = row_fill
            cell.alignment = alignment
            cell.border = CELL_BORDER
            if num_format:
                cell.number_format = num_format

        current_row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autofit_columns(ws, {1: 28, 2: 20, 15: 22, 16: 22, 17: 25, 18: 45})


# ---------------------------------------------------------------------------
# Sheet 4: Rationalisation Recommendation
# ---------------------------------------------------------------------------

def _build_rationalisation_sheet(
    ws: Any,
    portfolio: PortfolioAnalysis,
    rationalisation: RationalisationAnalysis | None,
) -> None:
    """Build Sheet 4: Rationalisation Recommendation with lossless evidence and zero synthetic artifacts."""
    ws.title = "Rationalisation Recommendation"

    headers = [
        "Workflow A",
        "Workflow B",
        "Business Area A",
        "Business Area B",
        "Opportunity Score",
        "Recommendation / Disposition",
        "Source Metadata Overlap %",
        "Target Metadata Overlap %",
        "Frequency Overlap %",
        "Logic Overlap %",
        "DAG Overlap %",
        "Shared Formulae",
        "Unique Workflow Functionality",
        "Justification",
    ]

    ws.row_dimensions[1].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_HEADER_CENTER if 5 <= col_idx <= 11 else ALIGN_HEADER
        cell.border = CELL_BORDER

    candidates = (rationalisation.candidates if rationalisation else []) or portfolio.rationalisation_candidates

    # Helper to convert normalized 0-1 metrics to standard Excel percentage floats
    def _to_pct_value(val: Any) -> float:
        try:
            f = float(val)
            if f > 1.0:
                return f / 100.0
            return max(0.0, f)
        except (ValueError, TypeError):
            return 0.0

    current_row = 2
    for idx, c in enumerate(candidates):
        row_fill = ZEBRA_FILL if idx % 2 == 1 else WHITE_FILL
        ws.row_dimensions[current_row].height = 50

        wf_a = c.workflow_names[0] if len(c.workflow_names) > 0 else (c.workflow_ids[0] if c.workflow_ids else "N/A")
        wf_b = c.workflow_names[1] if len(c.workflow_names) > 1 else (c.workflow_ids[1] if len(c.workflow_ids) > 1 else "Standalone")

        # Lookup business areas
        area_a = "Other / Unclassified"
        area_b = "Other / Unclassified"
        for w in portfolio.workflows:
            if w.filename == wf_a or w.workflow_id in c.workflow_ids[:1]:
                area_a = _normalize_tag(w.business_area_tag)
            if len(c.workflow_names) > 1 and (w.filename == wf_b or w.workflow_id in c.workflow_ids[1:2]):
                area_b = _normalize_tag(w.business_area_tag)

        opp_score = getattr(c, "opportunity_score", 0.0) or 0.0
        rec_type = c.recommendation_type or "REVIEW"

        # Sourced directly from deterministic_metrics (canonical rationalisation evidence)
        dm = getattr(c, "deterministic_metrics", None)
        if dm is not None:
            src_overlap = getattr(dm, "source_overlap", 0.0) or 0.0
            tgt_overlap = getattr(dm, "target_overlap", 0.0) or 0.0
            freq_overlap = getattr(dm, "frequency_overlap", 0.0) or 0.0
            trans_overlap = getattr(dm, "transformation_similarity", 0.0) or 0.0
            dag_overlap = getattr(dm, "dag_similarity", 0.0) or 0.0
        else:
            src_overlap = getattr(c, "source_overlap", 0.0) or 0.0
            tgt_overlap = getattr(c, "target_overlap", 0.0) or 0.0
            freq_overlap = getattr(c, "frequency_overlap", 0.0) or 0.0
            trans_overlap = getattr(c, "transformation_similarity", 0.0) or 0.0
            dag_overlap = getattr(c, "dag_similarity", getattr(c, "topology_similarity", 0.0)) or 0.0

        src_val = _to_pct_value(src_overlap)
        tgt_val = _to_pct_value(tgt_overlap)
        freq_val = _to_pct_value(freq_overlap)
        logic_val = _to_pct_value(trans_overlap)
        dag_val = _to_pct_value(dag_overlap)

        # Lossless shared formulae & unique functionality
        shared_logic_items = getattr(c, "shared_logic", []) or []
        shared_text = "\n".join(shared_logic_items) if shared_logic_items else "No shared formulae identified"

        unique_dict = getattr(c, "unique_functionality", {}) or {}
        unique_lines: list[str] = []
        for wf_name, ops in unique_dict.items():
            if ops:
                unique_lines.append(f"[{wf_name}]:")
                for op in ops:
                    unique_lines.append(f"  • {op}")
        unique_text = "\n".join(unique_lines) if unique_lines else "No distinguishing operations recorded"

        strategy = c.reasoning or getattr(c, "migration_disposition", "Assess for consolidation or shared macro extraction")

        row_values = [
            (wf_a, BOLD_BODY_FONT, ALIGN_WRAP_TOP, None),
            (wf_b, BOLD_BODY_FONT, ALIGN_WRAP_TOP, None),
            (area_a, BODY_FONT, ALIGN_WRAP_TOP, None),
            (area_b, BODY_FONT, ALIGN_WRAP_TOP, None),
            (opp_score, BOLD_BODY_FONT, ALIGN_CENTER_TOP, "0.0"),
            (rec_type, BOLD_BODY_FONT, ALIGN_CENTER_TOP, None),
            (src_val, BODY_FONT, ALIGN_CENTER_TOP, "0.0%"),
            (tgt_val, BODY_FONT, ALIGN_CENTER_TOP, "0.0%"),
            (freq_val, BODY_FONT, ALIGN_CENTER_TOP, "0.0%"),
            (logic_val, BODY_FONT, ALIGN_CENTER_TOP, "0.0%"),
            (dag_val, BODY_FONT, ALIGN_CENTER_TOP, "0.0%"),
            (shared_text, BODY_FONT, ALIGN_WRAP_TOP, None),
            (unique_text, BODY_FONT, ALIGN_WRAP_TOP, None),
            (strategy, BODY_FONT, ALIGN_WRAP_TOP, None),
        ]

        for col_idx, (val, font, alignment, num_format) in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font
            cell.fill = row_fill
            cell.alignment = alignment
            cell.border = CELL_BORDER
            if num_format:
                cell.number_format = num_format

        current_row += 1

    if not candidates:
        ws.cell(row=2, column=1, value="No rationalisation candidates identified for this portfolio.").font = MUTED_FONT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autofit_columns(ws, {1: 28, 2: 28, 3: 20, 4: 20, 5: 16, 6: 22, 7: 24, 8: 24, 9: 20, 10: 18, 11: 18, 12: 50, 13: 50, 14: 45})


# ---------------------------------------------------------------------------
# Master Workbook Generation Entrypoint
# ---------------------------------------------------------------------------

def generate_portfolio_excel(
    portfolio: PortfolioAnalysis,
    successful_results: dict[str, CanonicalAnalysisResult],
    rationalisation: RationalisationAnalysis | None,
    output_path: Path | str,
) -> None:
    """Generate a 4-sheet enterprise ETL Portfolio Overview Excel workbook.

    Args:
        portfolio: Canonical PortfolioAnalysis model.
        successful_results: Map of workflow_id -> CanonicalAnalysisResult.
        rationalisation: Optional pre-built or deterministic RationalisationAnalysis.
        output_path: Destination path for the .xlsx file.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # 1. Executive Summary
    ws_exec = wb.active
    _build_executive_summary_sheet(ws_exec, portfolio, rationalisation)

    # 2. Inventory
    ws_inv = wb.create_sheet()
    _build_portfolio_summary_sheet(ws_inv, portfolio, successful_results)

    # 3. Technical Inventory
    ws_tech = wb.create_sheet()
    _build_technical_inventory_sheet(ws_tech, portfolio, successful_results)

    # 4. Rationalisation Recommendation
    ws_rat = wb.create_sheet()
    _build_rationalisation_sheet(ws_rat, portfolio, rationalisation)

    wb.save(output_path)
