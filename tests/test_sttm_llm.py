"""Comprehensive test suite for LLM-Powered STTM, Mapping Authority Invariants, and Fallbacks."""

from __future__ import annotations

import json
from pathlib import Path
import tempfile
import openpyxl
import pytest

from awa.parser.xml_parser import parse_workflow
from awa.graph.builder import build_graph
from awa.analysis.workflow_analyzer import analyze_canonical
from awa.analysis.sttm_extractor import extract_sttm, build_sttm_evidence_context
from awa.analysis.sttm_validator import STTMValidator
from awa.llm.client import FakeLLMClient
from awa.llm.generator import LLMNarrativeGenerator
from awa.llm.cache import LLMNarrativeCache
from awa.generators.sttm_generator import generate_sttm_excel


class TestSTTMLLMIntegration:
    """Test suite verifying LLM STTM generation, deterministic validation, and invariants."""

    def test_unknown_wildcard_elimination_ftse(self):
        """Invariant 7 & Regression: *Unknown wildcard must NEVER appear as a mapping in FTSE 100."""
        wf_path = Path("FTSE 100.yxmd")
        if not wf_path.exists():
            pytest.skip("FTSE 100.yxmd not found")

        wf = parse_workflow(wf_path)
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        assert sttm.total_mappings > 0
        unknown_mappings = [
            m for m in sttm.mappings
            if "*unknown" in m.source_attribute.lower()
            or "*unknown" in m.target_attribute.lower()
            or m.source_attribute.startswith("*")
            or m.target_attribute.startswith("*")
        ]
        assert len(unknown_mappings) == 0, f"Found leaked *Unknown mappings: {unknown_mappings}"

        # Verify target attribute AdjustedClose exists
        adj_close = [m for m in sttm.mappings if m.target_attribute == "AdjustedClose"]
        assert len(adj_close) > 0
        assert adj_close[0].target_table == "FTSEData.tde"

    def test_actual_filename_precedence(self):
        """Invariant 4: Precedence requires actual configured file names/paths across all workflows."""
        # 1. Demo Claims
        wf_claims = parse_workflow(Path("Demo_Claims_Volume_Extract_reconstructed.yxmd"))
        g_claims = build_graph(wf_claims)
        ev_claims = build_sttm_evidence_context(wf_claims, g_claims)
        src_names = [s["dataset_name"] for s in ev_claims["source_datasets"]]
        tgt_names = [t["deliverable_name"] for t in ev_claims["target_deliverables"]]

        assert any("Claims_Volume_Extract_Demo.xlsx" in s for s in src_names)
        assert any("Policy_Master_Demo.xlsx" in s for s in src_names)
        assert any("Claim_Payments_Demo.xlsx" in s for s in src_names)
        assert any("Claim_Diary_Notes_Demo.xlsx" in s for s in src_names)
        assert any("Claims_Aging_Risk_Demo_Output.xlsx" in t for t in tgt_names)

        # 2. BBCFoodAggr
        wf_bbc = parse_workflow(Path("BBCFoodAggr.yxmd"))
        g_bbc = build_graph(wf_bbc)
        ev_bbc = build_sttm_evidence_context(wf_bbc, g_bbc)
        bbc_srcs = [s["dataset_name"] for s in ev_bbc["source_datasets"]]
        assert "BBCFood.yxdb" in bbc_srcs
        assert "BBCFoodChefs.yxdb" in bbc_srcs

        # 3. FTSE 100
        wf_ftse = parse_workflow(Path("FTSE 100.yxmd"))
        g_ftse = build_graph(wf_ftse)
        ev_ftse = build_sttm_evidence_context(wf_ftse, g_ftse)
        ftse_tgts = [t["deliverable_name"] for t in ev_ftse["target_deliverables"]]
        assert "FTSEData.tde" in ftse_tgts

    def test_bbcfood_aggr_browse_sink_distinction(self):
        """Constraint 2: BrowseV2 is an inspection sink, correctly producing 0 production STTM mappings."""
        wf = parse_workflow(Path("BBCFoodAggr.yxmd"))
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        assert sttm.total_mappings == 0

    def test_bbcfood_v2_regression(self):
        """Verify BBCFood v2 produces valid STTM mappings with zero wildcards."""
        wf = parse_workflow(Path("BBCFood v2.yxmd"))
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        assert sttm.total_mappings > 0
        for m in sttm.mappings:
            assert not m.source_attribute.startswith("*")
            assert not m.target_attribute.startswith("*")
        targets = {m.target_table for m in sttm.mappings}
        assert "BBCFood.yxdb" in targets

    def test_validator_rejects_hallucinations(self):
        """Mapping Authority Invariant: Validator strictly rejects fabricated entities."""
        wf = parse_workflow(Path("Demo_Claims_Volume_Extract_reconstructed.yxmd"))
        g = build_graph(wf)
        evidence = build_sttm_evidence_context(wf, g)
        validator = STTMValidator(evidence, g)

        valid_src = evidence["source_datasets"][0]["dataset_name"]
        valid_src_fld = evidence["source_datasets"][0]["fields"][0]
        valid_tgt = evidence["target_deliverables"][0]["deliverable_name"]
        valid_tgt_fld = evidence["target_deliverables"][0]["fields"][0]

        # 1. Hallucinated source dataset
        item1 = {
            "source_table": "Fake_NonExistent_Database.csv",
            "source_attribute": valid_src_fld,
            "target_table": valid_tgt,
            "target_attribute": valid_tgt_fld,
            "transformation": "Direct",
        }
        ok, reason = validator.validate_mapping(item1)
        assert not ok
        assert "not in authoritative source datasets" in reason

        # 2. Hallucinated source attribute
        item2 = {
            "source_table": valid_src,
            "source_attribute": "NonExistentColumn999",
            "target_table": valid_tgt,
            "target_attribute": valid_tgt_fld,
            "transformation": "Direct",
        }
        ok, reason = validator.validate_mapping(item2)
        assert not ok
        assert "not in known fields" in reason

        # 3. Hallucinated target deliverable
        item3 = {
            "source_table": valid_src,
            "source_attribute": valid_src_fld,
            "target_table": "Fabricated_Deliverable.xlsx",
            "target_attribute": valid_tgt_fld,
            "transformation": "Direct",
        }
        ok, reason = validator.validate_mapping(item3)
        assert not ok
        assert "not in authoritative target deliverables" in reason

        # 4. Hallucinated target attribute
        item4 = {
            "source_table": valid_src,
            "source_attribute": valid_src_fld,
            "target_table": valid_tgt,
            "target_attribute": "FakeOutputMetricXYZ",
            "transformation": "Direct",
        }
        ok, reason = validator.validate_mapping(item4)
        assert not ok
        assert "not in known fields" in reason

        # 5. Wildcard token rejection
        item5 = {
            "source_table": valid_src,
            "source_attribute": "*Unknown",
            "target_table": valid_tgt,
            "target_attribute": valid_tgt_fld,
            "transformation": "Direct",
        }
        ok, reason = validator.validate_mapping(item5)
        assert not ok
        assert "forbidden wildcard" in reason

        # 6. Disconnected DAG tools
        item6 = {
            "source_table": valid_src,
            "source_attribute": valid_src_fld,
            "target_table": valid_tgt,
            "target_attribute": valid_tgt_fld,
            "source_tool_id": 152,  # downstream sink
            "target_tool_id": 1,    # upstream source
            "transformation": "Direct",
        }
        ok, reason = validator.validate_mapping(item6)
        assert not ok
        assert "No DAG path connects" in reason

    def test_no_confidence_string_in_mappings(self):
        """User Constraint: Strictly no 'confidence' string in STTM mappings or dicts."""
        wf = parse_workflow(Path("Demo_Claims_Volume_Extract_reconstructed.yxmd"))
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        for m in sttm.mappings:
            d = m.to_dict()
            assert "confidence" not in d
            assert not hasattr(m, "confidence")

    def test_llm_generation_and_reconciliation(self):
        """Verify LLM generator enriches transformation logic and reconciles 100% target completeness."""
        wf = parse_workflow(Path("Demo_Claims_Volume_Extract_reconstructed.yxmd"))
        g = build_graph(wf)
        evidence = build_sttm_evidence_context(wf, g)
        baseline_total = len(evidence["deterministic_baseline"].mappings)

        # Mock LLM returns refined logic for 1 mapping, and hallucinates another
        target_name = evidence["target_deliverables"][0]["deliverable_name"]
        target_fld = evidence["target_deliverables"][0]["fields"][0]
        src_name = evidence["source_datasets"][0]["dataset_name"]
        src_fld = evidence["source_datasets"][0]["fields"][0]

        mock_payload = {
            "workflow_name": "Demo Claims",
            "mappings": [
                {
                    "source_table": src_name,
                    "source_attribute": src_fld,
                    "transformation": "Direct",
                    "transformation_logic": "AI-refined: Populates directly with full verification.",
                    "target_table": target_name,
                    "target_attribute": target_fld,
                    "source_tool_id": 1,
                    "target_tool_id": 17,
                },
                {
                    "source_table": "FakeSource",
                    "source_attribute": "FakeField",
                    "transformation": "Direct",
                    "transformation_logic": "Hallucinated entry",
                    "target_table": "FakeTarget",
                    "target_attribute": "FakeTargetField",
                },
            ],
        }

        fake_client = FakeLLMClient(default_response=json.dumps(mock_payload))
        generator = LLMNarrativeGenerator(client=fake_client, cache=LLMNarrativeCache())

        sttm_doc = generator.generate_sttm(wf, g, workflow_id="test_wf_llm")
        assert sttm_doc.total_mappings == baseline_total

        # Verify the valid LLM mapping was accepted and refined
        refined = [
            m for m in sttm_doc.mappings
            if m.target_table == target_name and m.target_attribute == target_fld
        ]
        assert len(refined) > 0
        assert refined[0].transformation_logic == "AI-refined: Populates directly with full verification."
        assert refined[0].source == "llm"

        # Verify hallucinated entry was rejected
        assert not any(m.source_table == "FakeSource" for m in sttm_doc.mappings)

    def test_llm_fallback_on_error(self):
        """Verify seamless fallback to deterministic baseline when LLM raises error or malformed JSON."""
        wf = parse_workflow(Path("Demo_Claims_Volume_Extract_reconstructed.yxmd"))
        g = build_graph(wf)

        # 1. Malformed JSON
        malformed_client = FakeLLMClient(default_response="Not a JSON string {invalid")
        gen1 = LLMNarrativeGenerator(client=malformed_client, cache=LLMNarrativeCache())
        sttm1 = gen1.generate_sttm(wf, g, workflow_id="test_malformed")
        assert sttm1.total_mappings == 31

        # 2. LLM Exception
        def raise_err(sys_p, usr_p):
            raise TimeoutError("LLM API timed out")

        error_client = FakeLLMClient(generator_fn=raise_err)
        gen2 = LLMNarrativeGenerator(client=error_client, cache=LLMNarrativeCache())
        sttm2 = gen2.generate_sttm(wf, g, workflow_id="test_error")
        assert sttm2.total_mappings == 31

    def test_xlsx_skeleton_unchanged(self):
        """Requirement 12 & 13: Existing STTM XLSX skeleton MUST remain unchanged."""
        wf = parse_workflow(Path("Demo_Claims_Volume_Extract_reconstructed.yxmd"))
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            tmp_path = Path(tf.name)

        try:
            generate_sttm_excel(sttm, tmp_path)
            wb = openpyxl.load_workbook(tmp_path)

            # Check sheet names
            assert wb.sheetnames == ["Source-to-Target Mapping", "STTM Summary"]

            # Sheet 1: Headers & Columns
            ws1 = wb["Source-to-Target Mapping"]
            expected_headers = [
                "Source Table",
                "Source Attribute",
                "Transformation",
                "Transformation Logic",
                "Target Table",
                "Target Attribute",
            ]
            actual_headers = [ws1.cell(row=1, column=c).value for c in range(1, 7)]
            assert actual_headers == expected_headers

            # Autofilter & freeze panes
            assert ws1.freeze_panes == "A2"
            assert ws1.auto_filter.ref.startswith("A1:F")

            # Sheet 2: Summary KPIs
            ws2 = wb["STTM Summary"]
            assert ws2.cell(row=2, column=2).value == "Source-to-Target Mapping Summary"
            assert ws2.cell(row=5, column=2).value == "Metric"
            assert ws2.cell(row=12, column=2).value == "Transformation Category"

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_download_sttm_endpoint_and_zip(self):
        """Verify the FastAPI STTM download endpoint and ZIP bundle inclusion using LLM pipeline."""
        import io
        import zipfile
        from starlette.testclient import TestClient
        from backend.app.main import app

        client = TestClient(app)
        wf_path = Path("Demo_Claims_Volume_Extract_reconstructed.yxmd")
        with open(wf_path, "rb") as f:
            resp = client.post("/api/upload", files={"file": ("Demo_Claims.yxmd", f, "application/xml")})

        assert resp.status_code == 200
        analysis_id = resp.json()["analysis_id"]

        sttm_resp = client.get(f"/api/download/{analysis_id}/sttm")
        assert sttm_resp.status_code == 200
        assert "spreadsheetml" in sttm_resp.headers["Content-Type"]

        # Validate that the downloaded workbook can be opened and has 2 sheets
        wb = openpyxl.load_workbook(io.BytesIO(sttm_resp.content))
        assert wb.sheetnames == ["Source-to-Target Mapping", "STTM Summary"]
        ws = wb["Source-to-Target Mapping"]
        assert ws.max_row == 32  # 1 header + 31 mappings

        zip_resp = client.get(f"/api/download/{analysis_id}/zip")
        assert zip_resp.status_code == 200

        with zipfile.ZipFile(io.BytesIO(zip_resp.content)) as zf:
            filenames = zf.namelist()
            assert any(f.endswith("_STTM.xlsx") for f in filenames)

