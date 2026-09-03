"""Comprehensive test suite for Portfolio Business Area Reconciliation and Executive Summary XLSX schema.

Covers:
1. Reconciled portfolio counts across all five business areas.
2. Underwriting workflows consuming Claims data remain Underwriting.
3. Claims & Risk workflows assigned correctly based on functional evidence.
4. Legal workflows assigned correctly.
5. Sales & Distribution workflows assigned correctly.
6. Genuinely unclassifiable workflows assigned to Other / Unclassified.
7. Canonical business_area_tag persisted for every workflow.
8. Missing/invalid tag fallback behaviour.
9. Executive Summary XLSX table column structure (exactly 4 columns).
10. Complete absence of % of Portfolio / % of Analysed Workflows in Executive Summary.
11. Complete absence of High Criticality Count in Executive Summary.
12. Agreement between Portfolio page card counts and XLSX Executive Summary counts.
13. Rendering of zero-workflow business areas in both UI and XLSX.
14. Total workflow count invariant (sum of 5 card counts == total analysed workflows).
15. Zero download-time LLM calls.
"""

from __future__ import annotations

import openpyxl
import pytest
from pathlib import Path

from awa.analysis.business_area_classifier import (
    ALLOWED_BUSINESS_AREAS,
    classify_business_area_deterministic,
    classify_business_function_deterministic,
)
from awa.analysis.portfolio_analyzer import (
    build_portfolio_analysis,
    ALL_PORTFOLIO_BUSINESS_AREAS,
    CONFIGURED_PORTFOLIO_BUSINESS_AREAS,
)
from awa.analysis.rationalisation_analyzer import build_rationalisation_analysis
from awa.generators.portfolio_xlsx_generator import generate_portfolio_excel
from awa.model.analysis_result import CanonicalAnalysisResult
from awa.model.business_summary import WorkflowBusinessSummary, BusinessRule
from awa.model.workflow import Workflow, WorkflowMetadata


from awa.analysis.workflow_analyzer import analyze_canonical


def _make_dummy_canonical_result(
    analysis_id: str,
    workflow_name: str,
    business_area_tag: str = "",
    business_function: str = "",
    business_purpose: str = "",
) -> CanonicalAnalysisResult:
    """Helper to construct a valid CanonicalAnalysisResult for testing."""
    res = analyze_canonical("Demo_Claims_Volume_Extract_reconstructed.yxmd", analysis_id=analysis_id)
    res.workflow.metadata.name = workflow_name
    if res.business_summary:
        res.business_summary.business_purpose = business_purpose
        res.business_summary.business_function = business_function
        res.business_summary.business_area_tag = business_area_tag
        if business_area_tag:
            res.business_summary.business_area_tag_source = "llm"
    return res


class TestBusinessAreaReconciliation:
    """Verification of the 15 required business-area reconciliation criteria."""

    def test_underwriting_consuming_claims_data_remains_underwriting(self):
        """Req 2: Underwriting workflows that consume Claims data must remain Underwriting

        when their primary function is underwriting decisioning, policy eligibility, pricing, rating, or risk assessment.
        """
        out_ev = [{
            "dataset": "Commercial_Policy_Pricing_Output.xlsx",
            "columns": ["PolicyNumber", "PremiumAmount", "RatingTier", "UnderwriterNotes"],
        }]
        res = classify_business_area_deterministic(
            out_ev,
            business_purpose="Processes historical claims volume data to calculate commercial policy pricing and rating factors for underwriting risk assessment.",
            workflow_name="Commercial_Policy_Pricing_Engine.yxmd",
            business_function="Policy pricing and premium rating calculation",
            input_sources=["Claims_Volume_Extract.xlsx", "Loss_History_Master.xlsx"],
        )
        assert res.business_area == "Underwriting"
        assert res.confidence in ("HIGH", "MEDIUM")

    def test_claims_and_risk_functional_assignment(self):
        """Req 3: Claims & Risk workflows assigned based on claims handling, adjudication, reserves, or fraud."""
        out_ev = [{
            "dataset": "Claims_Aging_Output.xlsx",
            "columns": ["ClaimID", "LossReserve", "AdjudicationStatus"],
        }]
        res = classify_business_area_deterministic(
            out_ev,
            business_purpose="Adjudicates open claims and calculates quarterly loss reserves for claims portfolio review.",
            workflow_name="Claims_Adjudication_and_Reserves.yxmd",
            business_function="Claims adjudication and loss reserve calculation",
            input_sources=["Policy_Master.xlsx"],
        )
        assert res.business_area == "Claims & Risk"
        assert res.confidence == "HIGH"

    def test_legal_functional_assignment(self):
        """Req 4: Legal workflows assigned based on statutory filings, regulatory compliance, or litigation."""
        out_ev = [{
            "dataset": "Compliance_Submission_Extract.xlsx",
            "columns": ["MatterID", "StatutoryFilingCode", "AuditPeriod"],
        }]
        res = classify_business_area_deterministic(
            out_ev,
            business_purpose="Generates regulatory compliance filings and statutory audit reports for state commissioners.",
            workflow_name="Statutory_Compliance_Reporting.yxmd",
            business_function="Regulatory compliance reporting and audit submission",
        )
        assert res.business_area == "Legal"
        assert res.confidence == "HIGH"

    def test_sales_and_distribution_functional_assignment(self):
        """Req 5: Sales & Distribution workflows assigned based on broker commissions or sales pipeline."""
        out_ev = [{
            "dataset": "Broker_Commission_Payout.xlsx",
            "columns": ["BrokerID", "CommissionAmount", "TerritoryCode", "QuotaAttainment"],
        }]
        res = classify_business_area_deterministic(
            out_ev,
            business_purpose="Calculates monthly broker commissions and tracks sales pipeline performance by territory.",
            workflow_name="Broker_Commission_and_Sales_Performance.yxmd",
            business_function="Broker commission calculation and territory sales performance",
        )
        assert res.business_area == "Sales & Distribution"
        assert res.confidence == "HIGH"

    def test_genuinely_unclassifiable_workflow_assignment(self):
        """Req 6: Genuinely unclassifiable workflows assigned to Other / Unclassified."""
        out_ev = [{
            "dataset": "recipe_ingredients_output.csv",
            "columns": ["RecipeID", "IngredientName", "Quantity"],
        }]
        res = classify_business_area_deterministic(
            out_ev,
            business_purpose="Parses culinary recipe text and aggregates ingredient quantities.",
            workflow_name="BBC_Food_Aggregator.yxmd",
            business_function="",
            input_sources=["bbc_recipes.xml"],
        )
        assert res.business_area == "Other / Unclassified"

    def test_missing_or_invalid_tag_triggers_deterministic_fallback(self):
        """Req 8: Missing, null, empty, or invalid tags must go through deterministic fallback."""
        # Analysis result with invalid/empty tag
        raw_res = _make_dummy_canonical_result(
            "wf_test_01",
            "Underwriting_Risk_Matrix.yxmd",
            business_area_tag="",
            business_purpose="Evaluates applicant risk scores and determines policy eligibility.",
            business_function="",
        )
        portfolio = build_portfolio_analysis([("Underwriting_Risk_Matrix.yxmd", "Underwriting_Risk_Matrix.yxmd", raw_res)])
        summary = portfolio.workflows[0]

        # Tag must be deterministically resolved to Underwriting, not left blank or UNCLASSIFIED
        assert summary.business_area_tag == "Underwriting"
        assert summary.business_area.business_area == "Underwriting"
        assert summary.business_area_tag_source == "deterministic_fallback"

    def test_five_area_portfolio_reconciliation_and_invariants(self, tmp_path):
        """Req 1, 7, 12, 13, 14: Full portfolio with all 5 areas verifying count reconciliation and invariants."""
        w1 = _make_dummy_canonical_result("wf1", "Claims_Fraud.yxmd", "Claims & Risk", "Claims fraud detection", "Detects suspicious claims")
        w2 = _make_dummy_canonical_result("wf2", "Legal_Audit.yxmd", "Legal", "Legal audit", "Tracks litigation")
        w3 = _make_dummy_canonical_result("wf3", "Policy_Rating.yxmd", "Underwriting", "Policy rating", "Calculates policy rating")
        w4 = _make_dummy_canonical_result("wf4", "Sales_Territory.yxmd", "Sales & Distribution", "Sales analysis", "Tracks territory sales")
        w5 = _make_dummy_canonical_result("wf5", "BBC_Food.yxmd", "Other / Unclassified", "", "Aggregates food recipes")
        w6 = _make_dummy_canonical_result("wf6", "Claims_Reserve.yxmd", "Claims & Risk", "Loss reserve", "Calculates claims reserves")
        w7 = _make_dummy_canonical_result("wf7", "Custom_Script.yxmd", "Other / Unclassified", "", "Runs generic XML parsing")

        raw_list = [
            ("Claims_Fraud.yxmd", "Claims_Fraud.yxmd", w1),
            ("Legal_Audit.yxmd", "Legal_Audit.yxmd", w2),
            ("Policy_Rating.yxmd", "Policy_Rating.yxmd", w3),
            ("Sales_Territory.yxmd", "Sales_Territory.yxmd", w4),
            ("BBC_Food.yxmd", "BBC_Food.yxmd", w5),
            ("Claims_Reserve.yxmd", "Claims_Reserve.yxmd", w6),
            ("Custom_Script.yxmd", "Custom_Script.yxmd", w7),
        ]

        portfolio = build_portfolio_analysis(raw_list)

        # Invariant: Authoritative total analysed count
        assert portfolio.workflow_count == 7
        assert portfolio.metrics.successful_workflows == 7

        # Invariant: 5 configured business area groups + 1 conditional Other / Unclassified group materialized
        assert len(portfolio.business_areas) == 6
        group_names = [g.business_area for g in portfolio.business_areas]
        assert set(group_names) == set(ALL_PORTFOLIO_BUSINESS_AREAS)
        other_group = next(g for g in portfolio.business_areas if g.business_area == "Other / Unclassified")
        assert other_group.workflow_count == 2
        assert len(other_group.workflows) == 2

        # Expected counts
        counts = portfolio.business_area_counts
        assert counts["Claims & Risk"] == 2
        assert counts["Legal"] == 1
        assert counts["Underwriting"] == 1
        assert counts["Sales & Distribution"] == 1
        assert counts["Other / Unclassified"] == 2

        # Invariant: Sum of all business-area card counts equals authoritative workflow count
        assert sum(counts.values()) == 7

        # Zero-count verification: what if an area has 0 workflows?
        zero_raw = [
            ("Claims_Fraud.yxmd", "Claims_Fraud.yxmd", w1),
            ("Sales_Territory.yxmd", "Sales_Territory.yxmd", w4),
        ]
        zero_portfolio = build_portfolio_analysis(zero_raw)
        assert len(zero_portfolio.business_areas) == 5
        assert zero_portfolio.business_area_counts["Legal"] == 0
        assert zero_portfolio.business_area_counts["Underwriting"] == 0
        assert zero_portfolio.business_area_counts["Other / Unclassified"] == 0
        assert sum(zero_portfolio.business_area_counts.values()) == 2

    def test_executive_summary_xlsx_columns_and_schema(self, tmp_path):
        """Req 9, 10, 11, 12, 15: Executive Summary table contains exactly 4 columns and excludes % of Portfolio and High Criticality Count."""
        w1 = _make_dummy_canonical_result("wf1", "Claims_Fraud.yxmd", "Claims & Risk", "Claims fraud detection", "Detects suspicious claims")
        w2 = _make_dummy_canonical_result("wf2", "Policy_Rating.yxmd", "Underwriting", "Policy rating", "Calculates policy pricing")
        w3 = _make_dummy_canonical_result("wf3", "BBC_Food.yxmd", "Other / Unclassified", "", "Parses recipe ingredients")

        raw_list = [
            ("Claims_Fraud.yxmd", "Claims_Fraud.yxmd", w1),
            ("Policy_Rating.yxmd", "Policy_Rating.yxmd", w2),
            ("BBC_Food.yxmd", "BBC_Food.yxmd", w3),
        ]
        portfolio = build_portfolio_analysis(raw_list)
        success_dict = {w.analysis_id: w for _, _, w in raw_list}
        rationalisation = build_rationalisation_analysis(portfolio, success_dict, use_llm=False)

        export_path = tmp_path / "ETL_Portfolio_Overview.xlsx"
        generate_portfolio_excel(portfolio, success_dict, rationalisation, export_path)

        wb = openpyxl.load_workbook(export_path)
        ws = wb["Executive Summary"]

        # 1. Header row verification (Row 9)
        headers = [ws.cell(row=9, column=c).value for c in range(1, 10) if ws.cell(row=9, column=c).value is not None]
        assert headers == ["Business Area", "Workflow Count", "Primary Function Focus", "Description"]
        assert len(headers) == 4

        # 2. Assert removed columns are COMPLETELY absent
        header_text = " ".join(headers)
        assert "%" not in header_text
        assert "Portfolio" not in header_text
        assert "Criticality" not in header_text
        assert "High" not in header_text

        # 3. Check all cells in columns 5 onwards in rows 9-15 are completely empty
        for r in range(9, 15):
            for c in range(5, 10):
                val = ws.cell(row=r, column=c).value
                assert val is None, f"Expected cell ({r}, {c}) to be None, got '{val}'"

        # 4. Check data rows for all 5 areas (Rows 10 to 14)
        area_rows: dict[str, int] = {}
        for r in range(10, 15):
            area_name = ws.cell(row=r, column=1).value
            count_val = ws.cell(row=r, column=2).value
            func_val = ws.cell(row=r, column=3).value
            desc_val = ws.cell(row=r, column=4).value
            assert area_name in ALL_PORTFOLIO_BUSINESS_AREAS
            assert isinstance(count_val, int)
            assert func_val is not None
            assert desc_val is not None
            area_rows[area_name] = count_val

        # Verify counts match portfolio.business_area_counts exactly
        assert area_rows["Claims & Risk"] == 1
        assert area_rows["Underwriting"] == 1
        assert area_rows["Other / Unclassified"] == 1
        assert area_rows["Legal"] == 0
        assert area_rows["Sales & Distribution"] == 0
        assert sum(area_rows.values()) == 3

    def test_llm_disabled_analyze_canonical_claims_volume(self):
        """Req 1, 11: When LLM is disabled/unavailable, analyze_canonical must invoke deterministic fallback

        and assign Claims & Risk with source='deterministic_fallback'.
        """
        res = analyze_canonical("Demo_Claims_Volume_Extract_reconstructed.yxmd", analysis_id="test_claims_canon")
        assert res.business_area_tag == "Claims & Risk"
        assert res.business_area_tag_source == "deterministic_fallback"
        assert res.business_summary is not None
        assert res.business_summary.business_area_tag == "Claims & Risk"
        assert res.business_summary.business_area_tag_source == "deterministic_fallback"
        assert "Claims adjudication" in res.business_function or "claims" in res.business_function.lower()
        assert len(res.business_summary.classification_evidence) > 0

    def test_llm_disabled_analyze_canonical_generic_utility(self):
        """Req 8: When LLM is disabled, a generic utility workflow must resolve to Other / Unclassified."""
        res = analyze_canonical("BBCFoodAggr.yxmd", analysis_id="test_bbcfood_canon")
        assert res.business_area_tag == "Other / Unclassified"
        assert res.business_area_tag_source == "deterministic_fallback"
        assert res.business_summary is not None
        assert res.business_summary.business_area_tag == "Other / Unclassified"

    def test_underwriting_decision_engine_named_workflow(self):
        """Req 2: LLM disabled/unavailable + workflow named Underwriting Decision Engine -> Underwriting."""
        res = classify_business_area_deterministic(
            [],
            workflow_name="Underwriting_Decision_Engine.yxmd",
            input_sources=["Claims_Volume_Master.xlsx"],
        )
        assert res.business_area == "Underwriting"
        assert res.confidence in ("HIGH", "MEDIUM")

    def test_premium_rating_using_claims_history(self):
        """Req 3: LLM disabled/unavailable + Premium Rating/Calculator using Claims history -> Underwriting."""
        res = classify_business_area_deterministic(
            [],
            business_purpose="Calculates commercial policy premium rating factors using past claims history.",
            workflow_name="Commercial_Premium_Rating_Calculator.yxmd",
            business_function="Policy premium calculation and rating",
            input_sources=["Claims_History.xlsx"],
        )
        assert res.business_area == "Underwriting"

    def test_claims_fraud_detection(self):
        """Req 4: LLM disabled/unavailable + Claims Fraud Detection -> Claims & Risk."""
        res = classify_business_area_deterministic(
            [],
            workflow_name="Claims_Fraud_Detection_Engine.yxmd",
            business_purpose="Identifies suspicious insurance claims using fraud detection algorithms.",
            business_function="Claims fraud investigation",
        )
        assert res.business_area == "Claims & Risk"

    def test_claims_reserve_calculator(self):
        """Req 5: LLM disabled/unavailable + Claims Reserve Calculator -> Claims & Risk."""
        res = classify_business_area_deterministic(
            [],
            workflow_name="Claims_Reserve_Calculator.yxmd",
            business_purpose="Calculates actuarial loss reserves for open insurance claims.",
            business_function="Loss reserve calculation",
        )
        assert res.business_area == "Claims & Risk"

    def test_sales_territory_analytics(self):
        """Req 6: LLM disabled/unavailable + Sales Territory Analytics -> Sales & Distribution."""
        res = classify_business_area_deterministic(
            [],
            workflow_name="Sales_Territory_Analytics.yxmd",
            business_purpose="Monitors product sales revenue and territory distribution quotas.",
            business_function="Sales territory and distribution performance",
        )
        assert res.business_area == "Sales & Distribution"

    def test_regulatory_compliance_reporting(self):
        """Req 7: LLM disabled/unavailable + Regulatory Compliance Reporting -> Legal."""
        res = classify_business_area_deterministic(
            [],
            workflow_name="Regulatory_Compliance_Reporting.yxmd",
            business_purpose="Prepares statutory compliance audit reports for regulatory filing.",
            business_function="Regulatory compliance and statutory audit reporting",
        )
        assert res.business_area == "Legal"

    def test_generate_business_purpose_receives_complete_evidence(self):
        """Req 9: Verify generator builds deterministic baseline from complete available evidence."""
        from awa.llm import get_default_generator
        gen = get_default_generator()
        # Build minimal test workflow
        w = Workflow(metadata=WorkflowMetadata(name="Claims_Loss_Analysis.yxmd", version="2021.4"))
        bs = WorkflowBusinessSummary(
            business_purpose="",
            one_line_purpose="",
            why_it_matters="",
            business_function="Claims loss reporting",
        )
        output_ev = [{"dataset": "Claims_Loss_Extract.xlsx", "columns": ["ClaimID", "LossAmount"]}]
        input_srcs = ["Policy_Master.xlsx"]
        res = gen.generate_business_purpose(
            w,
            bs,
            workflow_id="Claims_Loss_Analysis.yxmd",
            output_evidence=output_ev,
            input_sources=input_srcs,
        )
        assert res.business_area_tag == "Claims & Risk"
        assert res.source == "deterministic_fallback"
        assert any("Claims" in ev for ev in res.classification_evidence)

    def test_genuine_llm_other_unclassified_preserved(self):
        """Req 16: A genuine LLM classification of Other / Unclassified must NOT be reclassified."""
        res = _make_dummy_canonical_result(
            "wf_genuine_other",
            "Custom_ETL_Parser.yxmd",
            business_area_tag="Other / Unclassified",
            business_function="Custom technical data ETL",
            business_purpose="Custom technical XML ETL transformation",
        )
        # Mark provenance as LLM
        if res.business_summary:
            res.business_summary.business_area_tag_source = "llm"

        portfolio = build_portfolio_analysis([("Custom_ETL_Parser.yxmd", "Custom_ETL_Parser.yxmd", res)])
        summary = portfolio.workflows[0]

        assert summary.business_area_tag == "Other / Unclassified"
        assert summary.business_area_tag_source == "llm"

    def test_business_purpose_quality_rejects_reasoning_dumps(self):
        """Req 15, 16: Verify _is_clean_business_purpose rejects reasoning dumps and prompt leaks."""
        from awa.llm.generator import _is_clean_business_purpose

        # Disallowed: Tier breakdown / reasoning
        reasoning_sample = (
            "Tier 1 Workflow Name: matches Claims. Tier 2 Purpose Taxonomy indicates claims adjudication. "
            "Given these observations, we classify the workflow as Claims & Risk because the primary function is claims triage."
        )
        assert not _is_clean_business_purpose(reasoning_sample)

        # Disallowed: Chain of thought / analysis preamble
        cot_sample = "Let's analyze this step by step. First, the inputs show Policy_Master. Based on the facts provided, this is Underwriting."
        assert not _is_clean_business_purpose(cot_sample)

        # Disallowed: Multi-paragraph or markdown fences
        markdown_sample = "```json\n{\"business_purpose\": \"Valid purpose statement.\"}\n```"
        assert not _is_clean_business_purpose(markdown_sample)

        # Disallowed: Conversational lead-in
        lead_in = "Here is the business purpose: The workflow processes policy renewals and generates premium notices."
        assert not _is_clean_business_purpose(lead_in)

        # Allowed: Concise, polished business purpose
        clean_sample = (
            "Automates commercial underwriting risk evaluation and premium rating factor calculation for policy applications, "
            "evaluating applicant risk matrices to generate binding decisions and underwriter audit deliverables."
        )
        assert _is_clean_business_purpose(clean_sample)

    def test_llm_reasoning_rejection_falls_back_to_clean_purpose(self):
        """Req 16, 17: When LLM outputs reasoning in business_purpose, generator rejects it and uses clean fallback."""
        from unittest.mock import MagicMock
        from awa.llm.generator import LLMNarrativeGenerator
        from awa.llm.client import FakeLLMClient
        import json

        fake_client = FakeLLMClient(is_available=True)
        # LLM returns valid JSON but the business_purpose field contains a reasoning dump
        bad_llm_payload = {
            "business_purpose": "Tier 1 Workflow Name: Underwriting Decision. Tier 2 indicates underwriting. Rejected Claims & Risk because data is only supporting.",
            "business_function": "Underwriting Decisioning",
            "business_area_tag": "Underwriting",
        }
        fake_client.generate = MagicMock(return_value=json.dumps(bad_llm_payload))

        gen = LLMNarrativeGenerator(client=fake_client)
        w = Workflow(metadata=WorkflowMetadata(name="Underwriting_Decision_Engine.yxmd", version="2021.4"))
        bs = WorkflowBusinessSummary(
            business_purpose="Evaluates policy risk factors to determine insurance coverage eligibility.",
            one_line_purpose="",
            why_it_matters="",
            business_function="Underwriting Risk Assessment",
        )

        res = gen.generate_business_purpose(w, bs, workflow_id="Underwriting_Decision_Engine.yxmd")
        # Should reject the dirty LLM purpose and fall back to the clean deterministic purpose
        assert "Tier 1" not in res.business_purpose
        assert "Rejected" not in res.business_purpose
        assert res.source == "deterministic_fallback"
        assert res.business_area_tag == "Underwriting"
        assert res.business_purpose == bs.business_purpose  # Preserved clean deterministic purpose

    def test_cross_artifact_business_purpose_consistency(self):
        """Req 19: Portfolio overview, workflow details, and canonical result use the exact same Business Purpose."""
        res = analyze_canonical("Demo_Claims_Volume_Extract_reconstructed.yxmd", analysis_id="wf_consist_01")
        portfolio = build_portfolio_analysis([("Demo_Claims.yxmd", "Demo_Claims.yxmd", res)])

        port_wf = portfolio.workflows[0]
        # Canonical analysis result and portfolio workflow summary have identical purpose
        assert res.business_purpose == port_wf.business_purpose
        assert res.business_function == port_wf.business_function
        assert res.business_area_tag == port_wf.business_area_tag
        assert "Tier" not in port_wf.business_purpose

    def test_five_workflow_boundary_set(self):
        """Req 11: Explicit 5-workflow boundary test set (Underwriting, Claims, Underwriting Premium, Sales, Legal, Technical)."""
        # 1. Underwriting Decision Engine Application consuming Claims data
        r1 = classify_business_area_deterministic(
            workflow_name="Underwriting_Decision_Engine_Application.yxmd",
            business_purpose="Evaluates applicant risk rules to determine coverage eligibility and underwriter approval limits.",
            business_function="Underwriting decisioning and risk assessment",
            input_sources=["Claims_Submission_History.xlsx", "Applicant_Credit_Master.csv"],
            tool_configurations=['Formula Field: UnderwritingScore', 'Formula Expr: If [Score] > 700 Then "Approved" Else "Refer" EndIf'],
            container_titles=["Underwriting Eligibility Rules", "Risk Score Matrix"],
        )
        assert r1.business_area == "Underwriting"
        assert r1.confidence == "HIGH"

        # 2. Claims Fraud Detection
        r2 = classify_business_area_deterministic(
            workflow_name="Claims_Fraud_Detection.yxmd",
            business_purpose="Analyzes open claims for suspicious patterns, fraud indicators, and referral to special investigation unit.",
            business_function="Claims fraud detection and investigation prioritization",
            input_sources=["Claim_Payment_Detail.xlsx"],
            container_titles=["Fraud Scoring Engine"],
        )
        assert r2.business_area == "Claims & Risk"
        assert r2.confidence == "HIGH"

        # 3. Premium Calculator using Claims history
        r3 = classify_business_area_deterministic(
            workflow_name="Commercial_Policy_Premium_Calculator.yxmd",
            business_purpose="Calculates policyholder commercial premium rates and experience rating factors using prior claims loss history.",
            business_function="Policy pricing and premium rating calculation",
            input_sources=["Prior_Claims_Loss_History.xlsx", "Policy_Schedule.xlsx"],
            tool_configurations=['Formula Field: PremiumAmount', 'Formula Expr: [BaseRate] * [ExperienceModifier]'],
        )
        assert r3.business_area == "Underwriting"
        assert r3.confidence == "HIGH"

        # 4. Sales Territory Analytics
        r4 = classify_business_area_deterministic(
            workflow_name="Sales_Territory_Analytics.yxmd",
            business_purpose="Monitors quarterly distributor sales volume, territory quota attainment, and broker commissions.",
            business_function="Sales territory performance and distribution channel analytics",
            input_sources=["Distributor_Orders.xlsx", "Broker_Master.csv"],
            tool_configurations=['Formula Field: CommissionAmount', 'Formula Field: QuotaAttainment'],
            container_titles=["Territory Aggregation", "Broker Commission Engine"],
        )
        assert r4.business_area == "Sales & Distribution"
        assert r4.confidence == "HIGH"

        # 5. Regulatory Compliance Reporting
        r5 = classify_business_area_deterministic(
            workflow_name="Regulatory_Compliance_Reporting.yxmd",
            business_purpose="Generates annual statutory audit filings and compliance disclosure schedules for state insurance commissioners.",
            business_function="Regulatory compliance reporting and statutory filing",
            input_sources=["Statutory_Ledger.xlsx"],
            container_titles=["Regulatory Filings", "Statutory Compliance Audit"],
        )
        assert r5.business_area == "Legal"
        assert r5.confidence == "HIGH"

        # 6. Technical utility
        r6 = classify_business_area_deterministic(
            workflow_name="Generic_Xml_Parser.yxmd",
            business_purpose="",
            business_function="",
        )
        assert r6.business_area == "Other / Unclassified"
        assert r6.confidence == "UNCLASSIFIED"

    def test_deterministic_business_purpose_rich_paragraph_quality(self):
        """Req 7, 9: Deterministic Business Purpose is a polished ~40-75 word paragraph, not generic filler."""
        from awa.analysis.business_area_classifier import compose_deterministic_business_purpose
        from awa.model.business_summary import BusinessInput, BusinessOutput, BusinessRule

        w = Workflow(metadata=WorkflowMetadata(name="Commercial_Policy_Rating.yxmd", version="2021.4"))
        bs = WorkflowBusinessSummary(
            business_purpose="",
            one_line_purpose="",
            why_it_matters="",
            source_inputs=[
                BusinessInput(tool_id=1, name="Policy_Master.xlsx", raw_source="Policy_Master.xlsx", source_type="Excel"),
                BusinessInput(tool_id=2, name="Applicant_Credit.csv", raw_source="Applicant_Credit.csv", source_type="CSV"),
            ],
            business_outputs=[
                BusinessOutput(tool_id=3, name="Premium_Rating_Schedule.xlsx", raw_destination="Premium_Rating_Schedule.xlsx", destination_type="Excel"),
            ],
            business_rules=[
                BusinessRule(rule_name="Credit Tier Multiplier", category="Calculation", description="Applies tier multiplier"),
            ],
        )

        purpose = compose_deterministic_business_purpose(
            w,
            business_summary=bs,
            business_function="Policy pricing and premium rating calculation",
            business_area="Underwriting",
            workflow_name="Commercial_Policy_Rating.yxmd",
        )

        # Word count between 35 and 75
        words = purpose.split()
        assert 35 <= len(words) <= 75, f"Word count {len(words)} out of expected range: {purpose}"
        # Does not contain rejected generic filler
        assert "Automates policy pricing and premium rating calculation for Commercial_Policy_Rating.yxmd" not in purpose
        # Contains factual context
        assert "Policy Master" in purpose
        assert "Premium Rating Schedule" in purpose
        assert "calculation business rules" in purpose

    def test_evidence_hierarchy_extraction_from_workflow_structures(self):
        """Req 4, 6: extract_workflow_classification_evidence extracts containers, formulas, and annotations."""
        from awa.analysis.business_area_classifier import extract_workflow_classification_evidence
        from awa.model.container import ToolContainer
        from awa.model.tool import Tool, ToolConfiguration

        w = Workflow(
            metadata=WorkflowMetadata(name="Custom_App.yxmd", version="2021.4"),
            containers={
                1: ToolContainer(tool_id=1, caption="Underwriting Eligibility Filter"),
            },
            tools={
                2: Tool(
                    tool_id=2,
                    plugin="Formula",
                    tool_type="Formula",
                    name="Calculate Score",
                    position=None,
                    configuration=ToolConfiguration(
                        raw_xml="",
                        parsed={"formula_fields": [{"field": "UnderwritingScore", "expression": "[Score]*1.2"}]},
                    ),
                    annotation="Compute final underwriting score",
                ),
            },
        )
        ev = extract_workflow_classification_evidence(w)
        assert "Underwriting Eligibility Filter" in ev["container_titles"]
        assert any("UnderwritingScore" in f for f in ev["tool_configurations"])
        assert "Compute final underwriting score" in ev["tool_annotations"]

    def test_other_unclassified_conditional_materialization_zero_count(self):
        """When 0 workflows are unclassified, Other / Unclassified is NOT materialized in business_areas."""
        w1 = _make_dummy_canonical_result("wf1", "Claims.yxmd", "Claims & Risk", "Claims", "Claims handling")
        w2 = _make_dummy_canonical_result("wf2", "Actuarial.yxmd", "Actuarial", "Actuarial", "Loss reserving")
        portfolio = build_portfolio_analysis([
            ("Claims.yxmd", "Claims.yxmd", w1),
            ("Actuarial.yxmd", "Actuarial.yxmd", w2),
        ])

        assert portfolio.business_area_counts["Other / Unclassified"] == 0
        assert len(portfolio.business_areas) == 5
        group_names = [g.business_area for g in portfolio.business_areas]
        assert "Other / Unclassified" not in group_names
        assert set(group_names) == set(CONFIGURED_PORTFOLIO_BUSINESS_AREAS)

    def test_other_unclassified_conditional_materialization_positive_count(self):
        """When unclassifiable workflows exist, Other / Unclassified is conditionally materialized."""
        w1 = _make_dummy_canonical_result("wf1", "Claims.yxmd", "Claims & Risk", "Claims", "Claims handling")
        w_other1 = _make_dummy_canonical_result("wf_other1", "Generic_XML.yxmd", "Other / Unclassified", "", "Parses XML logs")
        w_other2 = _make_dummy_canonical_result("wf_other2", "DB_Migration.yxmd", "Other / Unclassified", "", "Test harness")

        portfolio = build_portfolio_analysis([
            ("Claims.yxmd", "Claims.yxmd", w1),
            ("Generic_XML.yxmd", "Generic_XML.yxmd", w_other1),
            ("DB_Migration.yxmd", "DB_Migration.yxmd", w_other2),
        ])

        assert portfolio.business_area_counts["Other / Unclassified"] == 2
        assert len(portfolio.business_areas) == 6
        group_names = [g.business_area for g in portfolio.business_areas]
        assert "Other / Unclassified" in group_names
        other_group = next(g for g in portfolio.business_areas if g.business_area == "Other / Unclassified")
        assert other_group.workflow_count == 2
        assert len(other_group.workflows) == 2
        assert {w.workflow_id for w in other_group.workflows} == {"wf_other1", "wf_other2"}


