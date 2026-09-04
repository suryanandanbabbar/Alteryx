"""Comprehensive automated test suite for ETL Portfolio Overview Excel export (.xlsx)."""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any
import openpyxl
import pytest
from starlette.testclient import TestClient

from awa.analysis.portfolio_analyzer import build_portfolio_analysis
from awa.analysis.rationalisation_analyzer import build_rationalisation_analysis
from awa.analysis.workflow_analyzer import analyze_canonical
from awa.generators.portfolio_xlsx_generator import generate_portfolio_excel
from awa.graph.builder import build_graph
from awa.model.analysis_result import CanonicalAnalysisResult
from awa.model.portfolio import PortfolioAnalysis, PortfolioWorkflowSummary
from awa.parser.xml_parser import parse_workflow
from backend.app.main import app
from backend.app.services.storage import get_storage


@pytest.fixture
def client():
    return TestClient(app)


@pytest.fixture
def sample_portfolio_and_results() -> tuple[PortfolioAnalysis, dict[str, CanonicalAnalysisResult]]:
    """Build a realistic multi-workflow portfolio with canonical analysis results."""
    # Workflow 1: Claims volume
    res1 = analyze_canonical("Demo_Claims_Volume_Extract_reconstructed.yxmd", analysis_id="wf_claims_01")
    res1.business_summary.business_purpose = "Processes and reconciles quarterly claims volume extract."
    res1.business_summary.business_function = "Claims Analytics & Reporting"
    res1.business_summary.business_area_tag = "Claims & Risk"

    # Workflow 2: FTSE
    res2 = analyze_canonical("FTSE 100.yxmd", analysis_id="wf_ftse_02")
    res2.business_summary.business_purpose = "Parses and aggregates FTSE market metrics."
    res2.business_summary.business_function = "Sales & Commercial Analytics"
    res2.business_summary.business_area_tag = "Sales & Distribution"

    raw_workflows = [
        ("Demo_Claims.yxmd", "Demo_Claims.yxmd", res1),
        ("FTSE_100.yxmd", "FTSE_100.yxmd", res2),
    ]

    portfolio = build_portfolio_analysis(
        raw_workflows,
        portfolio_name="Enterprise Test Estate",
        portfolio_id="test_port_001",
    )

    successful_results = {
        "wf_claims_01": res1,
        "wf_ftse_02": res2,
    }

    return portfolio, successful_results


class TestPortfolioXLSXExport:
    """Test suite for ETL Portfolio XLSX export generation and integrity."""

    def test_workbook_has_all_four_required_sheets(self, tmp_path, sample_portfolio_and_results):
        """Generated workbook must contain exactly the 4 required sheets in exact order."""
        portfolio, successful_results = sample_portfolio_and_results
        rationalisation = build_rationalisation_analysis(portfolio, successful_results, use_llm=False)

        export_file = tmp_path / "ETL_Portfolio_Overview.xlsx"
        generate_portfolio_excel(portfolio, successful_results, rationalisation, export_file)

        assert export_file.exists()
        wb = openpyxl.load_workbook(export_file)

        expected_sheets = [
            "Executive Summary",
            "Inventory",
            "Technical Inventory",
            "Rationalisation Recommendation",
        ]
        assert wb.sheetnames == expected_sheets
        assert "Workflow Features" not in wb.sheetnames
        assert "Portfolio Summary" not in wb.sheetnames
        assert "Rationalisation & Dependencies" not in wb.sheetnames

    def test_executive_summary_criticality_and_complexity_profile(self, tmp_path, sample_portfolio_and_results):
        """Executive Summary D4/D5 must display CRITICALITY PROFILE and E4/E5 COMPLEXITY PROFILE with dynamic H/M/L counts."""
        portfolio, successful_results = sample_portfolio_and_results
        rationalisation = build_rationalisation_analysis(portfolio, successful_results, use_llm=False)

        export_file = tmp_path / "ETL_Portfolio_Overview.xlsx"
        generate_portfolio_excel(portfolio, successful_results, rationalisation, export_file)

        wb = openpyxl.load_workbook(export_file)
        ws = wb["Executive Summary"]

        # Check title
        assert "Enterprise Test Estate" in str(ws["A1"].value)

        # Check D4 / D5 for CRITICALITY PROFILE
        assert ws["D4"].value == "CRITICALITY PROFILE"
        h_crit = sum(1 for w in portfolio.workflows if w.criticality_level == "HIGH")
        m_crit = sum(1 for w in portfolio.workflows if w.criticality_level == "MEDIUM")
        l_crit = sum(1 for w in portfolio.workflows if w.criticality_level == "LOW")
        assert ws["D5"].value == f"H:{h_crit} | M:{m_crit} | L:{l_crit}"

        # Check E4 / E5 for COMPLEXITY PROFILE
        assert ws["E4"].value == "COMPLEXITY PROFILE"
        h_comp = sum(1 for w in portfolio.workflows if w.complexity_level == "HIGH")
        m_comp = sum(1 for w in portfolio.workflows if w.complexity_level == "MEDIUM")
        l_comp = sum(1 for w in portfolio.workflows if w.complexity_level == "LOW")
        assert ws["E5"].value == f"H:{h_comp} | M:{m_comp} | L:{l_comp}"

    def test_criticality_profile_responds_dynamically(self, tmp_path, sample_portfolio_and_results):
        """Changing workflow criticality levels dynamically alters Executive Summary D5."""
        portfolio, successful_results = sample_portfolio_and_results
        # Explicitly set criticality levels
        portfolio.workflows[0].criticality_level = "HIGH"
        portfolio.workflows[1].criticality_level = "HIGH"

        export_file = tmp_path / "dynamic_crit.xlsx"
        generate_portfolio_excel(portfolio, successful_results, None, export_file)

        wb = openpyxl.load_workbook(export_file)
        ws = wb["Executive Summary"]
        assert ws["D4"].value == "CRITICALITY PROFILE"
        assert ws["D5"].value == "H:2 | M:0 | L:0"

        # Now change one to MEDIUM
        portfolio.workflows[0].criticality_level = "MEDIUM"
        generate_portfolio_excel(portfolio, successful_results, None, export_file)
        wb2 = openpyxl.load_workbook(export_file)
        assert wb2["Executive Summary"]["D5"].value == "H:1 | M:1 | L:0"

    def test_executive_summary_area_counts(self, tmp_path, sample_portfolio_and_results):
        """Executive Summary must display accurate KPIs and include ALL configured business areas (including 0-workflow areas)."""
        portfolio, successful_results = sample_portfolio_and_results
        rationalisation = build_rationalisation_analysis(portfolio, successful_results, use_llm=False)

        export_file = tmp_path / "ETL_Portfolio_Overview.xlsx"
        generate_portfolio_excel(portfolio, successful_results, rationalisation, export_file)

        wb = openpyxl.load_workbook(export_file)
        ws = wb["Executive Summary"]

        # Check business area table presence
        found_areas = set()
        for row in range(9, 20):
            area_val = ws.cell(row=row, column=1).value
            if area_val:
                found_areas.add(str(area_val).strip())

        # All 5 configured primary business areas + Other / Unclassified must exist
        assert "Claims & Risk" in found_areas
        assert "Legal" in found_areas
        assert "Underwriting" in found_areas
        assert "Sales & Distribution" in found_areas
        assert "Actuarial" in found_areas
        assert "Other / Unclassified" in found_areas

        # Zero workflow areas must be represented with count 0
        for row in range(9, 20):
            area_val = ws.cell(row=row, column=1).value
            if area_val in ("Legal", "Actuarial"):
                count_val = ws.cell(row=row, column=2).value
                assert count_val == 0

    def test_inventory_one_row_per_workflow(self, tmp_path, sample_portfolio_and_results):
        """Inventory must contain exactly one row per analysed workflow with canonical values, Last Run, and Frequency."""
        portfolio, successful_results = sample_portfolio_and_results
        # Add metadata for last_run and frequency to test population
        portfolio.workflows[0].last_run = "10 days ago"
        portfolio.workflows[0].frequency = "Daily"
        portfolio.workflows[1].last_run = "Not documented"
        portfolio.workflows[1].frequency = "Monthly"

        rationalisation = build_rationalisation_analysis(portfolio, successful_results, use_llm=False)

        export_file = tmp_path / "ETL_Portfolio_Overview.xlsx"
        generate_portfolio_excel(portfolio, successful_results, rationalisation, export_file)

        wb = openpyxl.load_workbook(export_file)
        ws = wb["Inventory"]

        # Check headers: Last Run | Frequency
        headers = [ws.cell(row=1, column=c).value for c in range(1, 24)]
        assert headers[20] == "Last Run"
        assert headers[21] == "Frequency"
        assert headers[22] == "Business Area Tag Source"

        # 2 workflows -> exactly rows 2 and 3
        wf_names = [ws.cell(row=r, column=1).value for r in (2, 3)]
        assert "Demo_Claims.yxmd" in wf_names
        assert "FTSE_100.yxmd" in wf_names

        # Row 4 must be empty
        assert ws.cell(row=4, column=1).value is None

        # Check canonical Business Purpose, Business Function, Last Run, and Frequency
        for r in (2, 3):
            name = ws.cell(row=r, column=1).value
            if name == "Demo_Claims.yxmd":
                assert ws.cell(row=r, column=2).value == "Claims & Risk"
                assert ws.cell(row=r, column=3).value == "Claims Analytics & Reporting"
                assert "reconciles quarterly claims" in ws.cell(row=r, column=4).value
                assert ws.cell(row=r, column=21).value == "10 days ago"
                assert ws.cell(row=r, column=22).value == "Daily"
            elif name == "FTSE_100.yxmd":
                assert ws.cell(row=r, column=2).value == "Sales & Distribution"
                assert ws.cell(row=r, column=3).value == "Sales & Commercial Analytics"
                assert "FTSE market metrics" in ws.cell(row=r, column=4).value
                assert ws.cell(row=r, column=21).value == "Not documented"
                assert ws.cell(row=r, column=22).value == "Monthly"

    def test_technical_inventory_deterministic_taxonomy(self, tmp_path, sample_portfolio_and_results):
        """Technical Inventory must deterministically count tool categories and flags without LLM guesses."""
        portfolio, successful_results = sample_portfolio_and_results

        export_file = tmp_path / "ETL_Portfolio_Overview.xlsx"
        generate_portfolio_excel(portfolio, successful_results, None, export_file)

        wb = openpyxl.load_workbook(export_file)
        ws = wb["Technical Inventory"]

        # Headers check
        headers = [ws.cell(row=1, column=c).value for c in range(1, 19)]
        assert "Formula Count" in headers
        assert "Join Count" in headers
        assert "Filter Count" in headers
        assert "Python Tool Present" in headers
        assert "SQL Present" in headers

        # Check Demo_Claims tool counts
        for r in (2, 3):
            name = ws.cell(row=r, column=1).value
            if name == "Demo_Claims.yxmd":
                tool_cnt = ws.cell(row=r, column=3).value
                assert tool_cnt > 0
                has_python = ws.cell(row=r, column=11).value
                assert has_python in ("Yes", "No")

    def test_rationalisation_recommendation_sheet_headers_and_metrics(self, tmp_path, sample_portfolio_and_results):
        """Rationalisation Recommendation sheet must contain exact renamed headers, no unwanted columns, and valid metrics."""
        portfolio, successful_results = sample_portfolio_and_results
        rationalisation = build_rationalisation_analysis(portfolio, successful_results, use_llm=False)

        export_file = tmp_path / "ETL_Portfolio_Overview.xlsx"
        generate_portfolio_excel(portfolio, successful_results, rationalisation, export_file)

        wb = openpyxl.load_workbook(export_file)
        ws = wb["Rationalisation Recommendation"]

        # Exact required headers in order
        expected_headers = [
            "Workflow A",
            "Workflow B",
            "Business Area A",
            "Business Area B",
            "Opportunity Score",
            "Recommendation / Disposition",
            "Source Metadata Overlap %",
            "Target Metadata Overlap %",
            "Frequency Overlap %",
            "Logic Overlap %",
            "DAG Overlap %",
            "Shared Formulae",
            "Unique Workflow Functionality",
            "Justification",
        ]
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(expected_headers) + 1)]
        assert headers == expected_headers

        # Verify unwanted columns are absent
        assert "Output Schema Overlap %" not in headers
        assert "Output Grain Alignment" not in headers
        assert "Output Grain Alignment %" not in headers

        # Check candidate rows if present
        for r in range(2, ws.max_row + 1):
            shared_logic = ws.cell(row=r, column=12).value
            if shared_logic:
                assert "Join on =" not in str(shared_logic)
                assert "Shared join key: =" not in str(shared_logic)

    def test_zero_download_time_llm_calls(self, tmp_path, sample_portfolio_and_results, monkeypatch):
        """Export generation must perform ZERO download-time LLM calls."""
        portfolio, successful_results = sample_portfolio_and_results

        from awa.llm.client import FakeLLMClient
        called = False

        def _boom(*args, **kwargs):
            nonlocal called
            called = True
            raise RuntimeError("LLM was called during export!")

        monkeypatch.setattr(FakeLLMClient, "generate", _boom)

        export_file = tmp_path / "zero_llm.xlsx"
        # Should generate without invoking LLM
        generate_portfolio_excel(portfolio, successful_results, None, export_file)

        assert export_file.exists()
        assert not called

    def test_partial_failed_workflow_handling(self, tmp_path, sample_portfolio_and_results):
        """Partial portfolio with a failed workflow must generate a valid, complete workbook."""
        portfolio, successful_results = sample_portfolio_and_results

        # Inject a failed workflow into the portfolio
        failed_wf = PortfolioWorkflowSummary(
            workflow_id="wf_failed_99",
            filename="Corrupt_File.yxmd",
            relative_path="Corrupt_File.yxmd",
            status="FAILED",
            error_message="XML Syntax Error: line 4 col 2",
        )
        portfolio.workflows.append(failed_wf)
        portfolio.metrics.total_workflows += 1
        portfolio.metrics.failed_workflows += 1

        export_file = tmp_path / "partial_portfolio.xlsx"
        generate_portfolio_excel(portfolio, successful_results, None, export_file)

        wb = openpyxl.load_workbook(export_file)
        ws_inv = wb["Inventory"]

        # Must have 3 rows now
        names = [ws_inv.cell(row=r, column=1).value for r in (2, 3, 4)]
        assert "Corrupt_File.yxmd" in names

        for r in (2, 3, 4):
            if ws_inv.cell(row=r, column=1).value == "Corrupt_File.yxmd":
                purpose = ws_inv.cell(row=r, column=4).value
                assert "Workflow analysis failed" in str(purpose)

    def test_api_endpoint_response_headers_and_status(self, client, sample_portfolio_and_results):
        """GET /api/portfolio/{portfolio_id}/export/xlsx must return 200, XLSX media type, and clean filename."""
        portfolio, successful_results = sample_portfolio_and_results
        storage = get_storage()
        storage.save_portfolio(portfolio)
        for wid, res in successful_results.items():
            storage.save(res)

        resp = client.get(f"/api/portfolio/{portfolio.portfolio_id}/export/xlsx")
        assert resp.status_code == 200
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp.headers["content-type"]
        assert 'attachment; filename="Enterprise_Test_Estate_Overview.xlsx"' in resp.headers["content-disposition"]

        # Validate that returned bytes form a valid openpyxl workbook
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert "Executive Summary" in wb.sheetnames
        assert "Inventory" in wb.sheetnames
        assert "Technical Inventory" in wb.sheetnames
        assert "Rationalisation Recommendation" in wb.sheetnames
        assert len(wb.sheetnames) == 4

    def test_api_endpoint_404_for_missing_portfolio(self, client):
        """GET /api/portfolio/missing_id/export/xlsx must return 404."""
        resp = client.get("/api/portfolio/non_existent_portfolio_id/export/xlsx")
        assert resp.status_code == 404
        assert "not found" in resp.json()["detail"]

    def test_actuarial_workflows_mapped_to_actuarial_bucket(self, tmp_path, sample_portfolio_and_results):
        """Actuarial workflows must be included dynamically in Actuarial area count and Inventory."""
        portfolio, successful_results = sample_portfolio_and_results

        # Add an Actuarial workflow
        act_wf = PortfolioWorkflowSummary(
            workflow_id="wf_actuarial_01",
            filename="Actuarial_Triangulation.yxmd",
            relative_path="Actuarial_Triangulation.yxmd",
            business_area_tag="Actuarial",
            business_function="Actuarial Reserving",
            business_purpose="Performs loss development triangulation and IBNR reserving.",
            last_run="2024-03-01",
            frequency="Quarterly",
            status="SUCCESS",
        )
        portfolio.workflows.append(act_wf)
        portfolio.metrics.total_workflows += 1
        portfolio.metrics.successful_workflows += 1

        export_file = tmp_path / "actuarial_test.xlsx"
        generate_portfolio_excel(portfolio, successful_results, None, export_file)

        wb = openpyxl.load_workbook(export_file)
        ws_exec = wb["Executive Summary"]
        ws_inv = wb["Inventory"]

        # In Executive Summary, Actuarial count must be 1
        actuarial_found = False
        for r in range(9, 20):
            if ws_exec.cell(row=r, column=1).value == "Actuarial":
                actuarial_found = True
                assert ws_exec.cell(row=r, column=2).value == 1
                assert ws_exec.cell(row=r, column=3).value == "Actuarial Reserving"
        assert actuarial_found

        # In Inventory, Actuarial_Triangulation.yxmd must have Business Area = Actuarial, Last Run = 2024-03-01, Frequency = Quarterly
        inv_names = [ws_inv.cell(row=r, column=1).value for r in range(2, 5)]
        assert "Actuarial_Triangulation.yxmd" in inv_names
        for r in range(2, 5):
            if ws_inv.cell(row=r, column=1).value == "Actuarial_Triangulation.yxmd":
                assert ws_inv.cell(row=r, column=2).value == "Actuarial"
                assert ws_inv.cell(row=r, column=21).value == "2024-03-01"
                assert ws_inv.cell(row=r, column=22).value == "Quarterly"

    def test_rationalisation_similarity_metrics_populated_accurately(self, tmp_path, sample_portfolio_and_results):
        """Rationalisation Recommendation columns must carry exact deterministic similarity values with 0.0% formatting."""
        portfolio, successful_results = sample_portfolio_and_results

        from awa.model.portfolio import RationalisationAnalysis, RationalisationCandidate, DeterministicMetrics

        cand = RationalisationCandidate(
            workflow_ids=["wf_claims_01", "wf_ftse_02"],
            workflow_names=["Demo_Claims.yxmd", "FTSE_100.yxmd"],
            recommendation_type="CONSOLIDATE",
            opportunity_score=85.5,
            deterministic_metrics=DeterministicMetrics(
                source_overlap=1.0,
                target_overlap=0.625,
                frequency_overlap=1.0,
                transformation_similarity=0.45,
                dag_similarity=0.75,
            ),
            shared_logic=["Formula: [Amount] * 1.2", "Filter: [Active] = 1"],
            unique_functionality={"Demo_Claims.yxmd": ["Summarize claims"]},
            reasoning="Consolidation candidate due to identical sources and matching frequency.",
        )

        rat = RationalisationAnalysis(
            portfolio_id=portfolio.portfolio_id,
            candidates=[cand],
        )

        export_file = tmp_path / "rat_metrics_test.xlsx"
        generate_portfolio_excel(portfolio, successful_results, rat, export_file)

        wb = openpyxl.load_workbook(export_file)
        ws = wb["Rationalisation Recommendation"]

        assert ws.cell(row=2, column=1).value == "Demo_Claims.yxmd"
        assert ws.cell(row=2, column=2).value == "FTSE_100.yxmd"
        assert ws.cell(row=2, column=5).value == 85.5
        assert ws.cell(row=2, column=6).value == "CONSOLIDATE"

        # Col 7: Source Metadata Overlap %
        assert ws.cell(row=2, column=7).value == 1.0
        assert ws.cell(row=2, column=7).number_format == "0.0%"

        # Col 8: Target Metadata Overlap %
        assert ws.cell(row=2, column=8).value == 0.625
        assert ws.cell(row=2, column=8).number_format == "0.0%"

        # Col 9: Frequency Overlap %
        assert ws.cell(row=2, column=9).value == 1.0
        assert ws.cell(row=2, column=9).number_format == "0.0%"

        # Col 10: Logic Overlap %
        assert ws.cell(row=2, column=10).value == 0.45
        assert ws.cell(row=2, column=10).number_format == "0.0%"

        # Col 11: DAG Overlap %
        assert ws.cell(row=2, column=11).value == 0.75
        assert ws.cell(row=2, column=11).number_format == "0.0%"

        # Col 12: Shared Formulae
        shared_val = ws.cell(row=2, column=12).value
        assert "Formula: [Amount] * 1.2" in shared_val
        assert "Filter: [Active] = 1" in shared_val

