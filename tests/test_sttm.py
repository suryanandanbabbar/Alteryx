"""Comprehensive tests for Generic Source-to-Target Mapping (STTM).

Validates that STTM extraction is 100% generic, domain-agnostic, deterministic,
and production-ready across multiple structurally distinct workflows.
"""

import inspect
import io
from pathlib import Path
import openpyxl
import pytest
from fastapi.testclient import TestClient

from awa.analysis.workflow_analyzer import analyze_canonical
from awa.analysis.sttm_extractor import extract_sttm
from awa.generators.sttm_generator import generate_sttm_excel
from awa.parser.xml_parser import parse_workflow
from awa.graph.builder import build_graph, execution_order
from awa.analysis.business_intelligence import generate_business_summary
from backend.app.main import app
import awa.analysis.sttm_extractor as sttm_module


class TestGenericSTTM:
    """Test generic STTM lineage extraction across various transformation patterns."""

    def test_no_hardcoded_domain_constants_in_sttm_code(self):
        """Static audit ensuring zero domain-specific or workflow-specific constants in sttm_extractor.py."""
        source_code = inspect.getsource(sttm_module)

        forbidden_terms = [
            "Demo_Claims",
            "Claims Volume",
            "Policy Master",
            "Claim Payments",
            "Claim Diary Notes",
            "Claim Number",
            "Payment Amount",
            "Total Paid",
            "Aging Bucket",
            "Last Activity Date",
            "Litigation Flag",
            "Preclaim",
            "Active_Pending",
            "Approved",
            "Stable_and_Mature",
            "Product Type",
        ]

        for term in forbidden_terms:
            assert term.lower() not in source_code.lower(), f"Forbidden hard-coded term '{term}' found in sttm_extractor.py!"

    def test_direct_mapping_generic(self, tmp_path: Path):
        """TEST 1: Direct pass-through mapping (Input.A -> Output.A)."""
        xml = """<?xml version="1.0"?>
<AlteryxDocument yxmdVer="2024.1">
  <Nodes>
    <Node ToolID="1"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput" /><Properties><Configuration><File>accounts.csv</File></Configuration></Properties></Node>
    <Node ToolID="2"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileOutput.DbFileOutput" /><Properties><Configuration><File>accounts_backup.csv</File></Configuration></Properties></Node>
  </Nodes>
  <Connections><Connection><Origin ToolID="1" Connection="Output" /><Destination ToolID="2" Connection="Input" /></Connection></Connections>
</AlteryxDocument>"""
        wf_file = tmp_path / "direct_test.yxmd"
        wf_file.write_text(xml, encoding="utf-8")
        wf = parse_workflow(wf_file)
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        assert sttm.total_mappings >= 1
        m = sttm.mappings[0]
        assert m.transformation == "Direct"
        assert "direct" in m.transformation_logic.lower()

    def test_rename_mapping_generic(self, tmp_path: Path):
        """TEST 2 & 10: Select projection & renaming (Input.customer_id -> Target.client_id)."""
        xml = """<?xml version="1.0"?>
<AlteryxDocument yxmdVer="2024.1">
  <Nodes>
    <Node ToolID="1"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput" /><Properties><Configuration><File>users.csv</File></Configuration></Properties></Node>
    <Node ToolID="2"><GuiSettings Plugin="AlteryxBasePluginsGui.AlteryxSelect.AlteryxSelect" /><Properties><Configuration>
      <SelectFields><SelectField field="user_id" selected="True" rename="account_id" /></SelectFields>
    </Configuration></Properties></Node>
    <Node ToolID="3"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileOutput.DbFileOutput" /><Properties><Configuration><File>dim_account.csv</File></Configuration></Properties></Node>
  </Nodes>
  <Connections>
    <Connection><Origin ToolID="1" Connection="Output" /><Destination ToolID="2" Connection="Input" /></Connection>
    <Connection><Origin ToolID="2" Connection="Output" /><Destination ToolID="3" Connection="Input" /></Connection>
  </Connections>
</AlteryxDocument>"""
        wf_file = tmp_path / "rename_test.yxmd"
        wf_file.write_text(xml, encoding="utf-8")
        wf = parse_workflow(wf_file)
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        target_map = {m.target_attribute: m for m in sttm.mappings}
        assert "account_id" in target_map
        assert target_map["account_id"].transformation == "Rename"
        assert target_map["account_id"].source_attribute == "user_id"

    def test_formula_arithmetic_multi_origin(self, tmp_path: Path):
        """TEST 3 & 12: Formula derived field with multiple source origins ([Qty] * [Price] -> [GrossAmount])."""
        xml = """<?xml version="1.0"?>
<AlteryxDocument yxmdVer="2024.1">
  <Nodes>
    <Node ToolID="1"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput" /><Properties><Configuration><File>orders.csv</File></Configuration></Properties></Node>
    <Node ToolID="2"><GuiSettings Plugin="AlteryxBasePluginsGui.Formula.Formula" /><Properties><Configuration>
      <FormulaFields><FormulaField field="GrossAmount" expression="[Qty] * [Price]" type="Double" size="8" /></FormulaFields>
    </Configuration></Properties></Node>
    <Node ToolID="3"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileOutput.DbFileOutput" /><Properties><Configuration><File>order_totals.csv</File></Configuration></Properties></Node>
  </Nodes>
  <Connections>
    <Connection><Origin ToolID="1" Connection="Output" /><Destination ToolID="2" Connection="Input" /></Connection>
    <Connection><Origin ToolID="2" Connection="Output" /><Destination ToolID="3" Connection="Input" /></Connection>
  </Connections>
</AlteryxDocument>"""
        wf_file = tmp_path / "formula_test.yxmd"
        wf_file.write_text(xml, encoding="utf-8")
        wf = parse_workflow(wf_file)
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        gross_mappings = [m for m in sttm.mappings if m.target_attribute == "GrossAmount"]
        assert len(gross_mappings) == 2, f"Expected 2 source origins (Qty, Price), got {len(gross_mappings)}"
        src_attrs = {m.source_attribute for m in gross_mappings}
        assert src_attrs == {"Qty", "Price"}
        for m in gross_mappings:
            assert m.transformation == "Derived Calculation"

    def test_join_lineage_generic(self, tmp_path: Path):
        """TEST 4: Relational Join with Left payload, Right payload, and match keys."""
        xml = """<?xml version="1.0"?>
<AlteryxDocument yxmdVer="2024.1">
  <Nodes>
    <Node ToolID="1"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput" /><Properties><Configuration><File>orders.csv</File></Configuration></Properties></Node>
    <Node ToolID="2"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput" /><Properties><Configuration><File>customers.csv</File></Configuration></Properties></Node>
    <Node ToolID="3"><GuiSettings Plugin="AlteryxBasePluginsGui.Join.Join" /><Properties><Configuration>
      <JoinInfo connection="Left"><Field field="cust_id" /></JoinInfo>
      <JoinInfo connection="Right"><Field field="cust_id" /></JoinInfo>
    </Configuration></Properties></Node>
    <Node ToolID="4"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileOutput.DbFileOutput" /><Properties><Configuration><File>enriched_orders.csv</File></Configuration></Properties></Node>
  </Nodes>
  <Connections>
    <Connection><Origin ToolID="1" Connection="Output" /><Destination ToolID="3" Connection="Left" /></Connection>
    <Connection><Origin ToolID="2" Connection="Output" /><Destination ToolID="3" Connection="Right" /></Connection>
    <Connection><Origin ToolID="3" Connection="Join" /><Destination ToolID="4" Connection="Input" /></Connection>
  </Connections>
</AlteryxDocument>"""
        wf_file = tmp_path / "join_test.yxmd"
        wf_file.write_text(xml, encoding="utf-8")
        wf = parse_workflow(wf_file)
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        assert sttm.total_mappings >= 1
        # Check join key origin comes from primary Left input
        cust_id_mappings = [m for m in sttm.mappings if m.target_attribute == "cust_id"]
        assert len(cust_id_mappings) == 1
        assert "order" in cust_id_mappings[0].source_table.lower()

    def test_aggregation_and_count_distinct_generic(self, tmp_path: Path):
        """TEST 5 & 6: Generic Aggregation (SUM, COUNTDISTINCT, GROUPBY)."""
        xml = """<?xml version="1.0"?>
<AlteryxDocument yxmdVer="2024.1">
  <Nodes>
    <Node ToolID="1"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput" /><Properties><Configuration><File>sales.csv</File></Configuration></Properties></Node>
    <Node ToolID="2"><GuiSettings Plugin="AlteryxBasePluginsGui.Summarize.Summarize" /><Properties><Configuration>
      <SummarizeFields>
        <SummarizeField field="region" action="GroupBy" rename="Region" />
        <SummarizeField field="amount" action="Sum" rename="TotalRevenue" />
        <SummarizeField field="buyer_id" action="CountDistinct" rename="UniqueBuyers" />
      </SummarizeFields>
    </Configuration></Properties></Node>
    <Node ToolID="3"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileOutput.DbFileOutput" /><Properties><Configuration><File>sales_by_region.csv</File></Configuration></Properties></Node>
  </Nodes>
  <Connections>
    <Connection><Origin ToolID="1" Connection="Output" /><Destination ToolID="2" Connection="Input" /></Connection>
    <Connection><Origin ToolID="2" Connection="Output" /><Destination ToolID="3" Connection="Input" /></Connection>
  </Connections>
</AlteryxDocument>"""
        wf_file = tmp_path / "summarize_test.yxmd"
        wf_file.write_text(xml, encoding="utf-8")
        wf = parse_workflow(wf_file)
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        target_map = {m.target_attribute: m for m in sttm.mappings}
        assert "Region" in target_map
        assert target_map["Region"].transformation == "Aggregation"
        assert "group" in target_map["Region"].transformation_logic.lower()

        assert "TotalRevenue" in target_map
        assert target_map["TotalRevenue"].transformation == "Aggregation"
        assert target_map["TotalRevenue"].source_attribute == "amount"
        assert "SUM" in target_map["TotalRevenue"].transformation_logic

        assert "UniqueBuyers" in target_map
        assert target_map["UniqueBuyers"].transformation == "Aggregation"
        assert target_map["UniqueBuyers"].source_attribute == "buyer_id"
        assert "COUNTDISTINCT" in target_map["UniqueBuyers"].transformation_logic

    def test_dynamic_crosstab_pivot_generic(self, tmp_path: Path):
        """TEST 7: Dynamic CrossTab without hardcoded column names (e.g. Sales by Quarters: Q1, Q2, Q3, Q4)."""
        xml = """<?xml version="1.0"?>
<AlteryxDocument yxmdVer="2024.1">
  <Nodes>
    <Node ToolID="1"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput" /><Properties><Configuration><File>quarterly_sales.csv</File></Configuration></Properties></Node>
    <Node ToolID="2"><GuiSettings Plugin="AlteryxBasePluginsGui.CrossTab.CrossTab" /><Properties><Configuration>
      <GroupFields><Field field="Department" /></GroupFields>
      <HeaderField field="QuarterName" />
      <DataField field="SalesVolume" />
      <Method>Sum</Method>
    </Configuration></Properties></Node>
    <Node ToolID="3"><GuiSettings Plugin="AlteryxBasePluginsGui.AlteryxSelect.AlteryxSelect" /><Properties><Configuration>
      <SelectFields>
        <SelectField field="Department" selected="True" />
        <SelectField field="Q1" selected="True" />
        <SelectField field="Q2" selected="True" />
        <SelectField field="Q3" selected="True" />
        <SelectField field="Q4" selected="True" />
      </SelectFields>
    </Configuration></Properties></Node>
    <Node ToolID="4"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileOutput.DbFileOutput" /><Properties><Configuration><File>dept_quarterly_matrix.csv</File></Configuration></Properties></Node>
  </Nodes>
  <Connections>
    <Connection><Origin ToolID="1" Connection="Output" /><Destination ToolID="2" Connection="Input" /></Connection>
    <Connection><Origin ToolID="2" Connection="Output" /><Destination ToolID="3" Connection="Input" /></Connection>
    <Connection><Origin ToolID="3" Connection="Output" /><Destination ToolID="4" Connection="Input" /></Connection>
  </Connections>
</AlteryxDocument>"""
        wf_file = tmp_path / "crosstab_test.yxmd"
        wf_file.write_text(xml, encoding="utf-8")
        wf = parse_workflow(wf_file)
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        # Verify Department grouping key
        dept_m = [m for m in sttm.mappings if m.target_attribute == "Department"]
        assert len(dept_m) == 1
        assert dept_m[0].transformation == "Pivot / Reshape"

        # Verify Q1, Q2, Q3, Q4 each receive DUAL dependencies: SalesVolume measure and QuarterName category
        for q in ["Q1", "Q2", "Q3", "Q4"]:
            q_mappings = [m for m in sttm.mappings if m.target_attribute == q]
            assert len(q_mappings) == 2, f"Expected 2 dual dependencies for {q}, got {len(q_mappings)}"
            src_attrs = {m.source_attribute for m in q_mappings}
            assert "SalesVolume" in src_attrs
            assert "QuarterName" in src_attrs

    def test_union_multi_source_generic(self, tmp_path: Path):
        """TEST 8: Union combining records from two independent sources."""
        xml = """<?xml version="1.0"?>
<AlteryxDocument yxmdVer="2024.1">
  <Nodes>
    <Node ToolID="1"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput" /><Properties><Configuration><File>north_region.csv</File></Configuration></Properties></Node>
    <Node ToolID="2"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput" /><Properties><Configuration><File>south_region.csv</File></Configuration></Properties></Node>
    <Node ToolID="3"><GuiSettings Plugin="AlteryxBasePluginsGui.Union.Union" /><Properties><Configuration><Mode>ByName</Mode></Configuration></Properties></Node>
    <Node ToolID="4"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileOutput.DbFileOutput" /><Properties><Configuration><File>all_regions.csv</File></Configuration></Properties></Node>
  </Nodes>
  <Connections>
    <Connection><Origin ToolID="1" Connection="Output" /><Destination ToolID="3" Connection="Input" /></Connection>
    <Connection><Origin ToolID="2" Connection="Output" /><Destination ToolID="3" Connection="Input" /></Connection>
    <Connection><Origin ToolID="3" Connection="Output" /><Destination ToolID="4" Connection="Input" /></Connection>
  </Connections>
</AlteryxDocument>"""
        wf_file = tmp_path / "union_test.yxmd"
        wf_file.write_text(xml, encoding="utf-8")
        wf = parse_workflow(wf_file)
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        # Both North and South sources must contribute
        sources = {m.source_table for m in sttm.mappings}
        assert any("north" in s.lower() for s in sources)
        assert any("south" in s.lower() for s in sources)

    def test_filter_pass_through_generic(self, tmp_path: Path):
        """TEST 9: Filter passes through records without mutating field-level provenance."""
        xml = """<?xml version="1.0"?>
<AlteryxDocument yxmdVer="2024.1">
  <Nodes>
    <Node ToolID="1"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput" /><Properties><Configuration><File>inventory.csv</File></Configuration></Properties></Node>
    <Node ToolID="2"><GuiSettings Plugin="AlteryxBasePluginsGui.Filter.Filter" /><Properties><Configuration><Expression>[Stock] &gt; 0</Expression></Configuration></Properties></Node>
    <Node ToolID="3"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileOutput.DbFileOutput" /><Properties><Configuration><File>active_stock.csv</File></Configuration></Properties></Node>
  </Nodes>
  <Connections>
    <Connection><Origin ToolID="1" Connection="Output" /><Destination ToolID="2" Connection="Input" /></Connection>
    <Connection><Origin ToolID="2" Connection="True" /><Destination ToolID="3" Connection="Input" /></Connection>
  </Connections>
</AlteryxDocument>"""
        wf_file = tmp_path / "filter_test.yxmd"
        wf_file.write_text(xml, encoding="utf-8")
        wf = parse_workflow(wf_file)
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        assert sttm.total_mappings >= 1
        m = sttm.mappings[0]
        assert m.transformation == "Direct"

    def test_multi_stage_transitive_pipeline(self, tmp_path: Path):
        """TEST 11: Multi-stage pipeline: Input -> Formula -> Select -> Join -> Summarize -> Output."""
        xml = """<?xml version="1.0"?>
<AlteryxDocument yxmdVer="2024.1">
  <Nodes>
    <Node ToolID="1"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput" /><Properties><Configuration><File>transactions.csv</File></Configuration></Properties></Node>
    <Node ToolID="2"><GuiSettings Plugin="AlteryxBasePluginsGui.Formula.Formula" /><Properties><Configuration>
      <FormulaFields><FormulaField field="NetVal" expression="[Gross] - [Discount]" type="Double" size="8" /></FormulaFields>
    </Configuration></Properties></Node>
    <Node ToolID="3"><GuiSettings Plugin="AlteryxBasePluginsGui.AlteryxSelect.AlteryxSelect" /><Properties><Configuration>
      <SelectFields><SelectField field="NetVal" selected="True" rename="AdjustedNet" /></SelectFields>
    </Configuration></Properties></Node>
    <Node ToolID="4"><GuiSettings Plugin="AlteryxBasePluginsGui.Summarize.Summarize" /><Properties><Configuration>
      <SummarizeFields><SummarizeField field="AdjustedNet" action="Sum" rename="TotalAdjusted" /></SummarizeFields>
    </Configuration></Properties></Node>
    <Node ToolID="5"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileOutput.DbFileOutput" /><Properties><Configuration><File>financial_summary.csv</File></Configuration></Properties></Node>
  </Nodes>
  <Connections>
    <Connection><Origin ToolID="1" Connection="Output" /><Destination ToolID="2" Connection="Input" /></Connection>
    <Connection><Origin ToolID="2" Connection="Output" /><Destination ToolID="3" Connection="Input" /></Connection>
    <Connection><Origin ToolID="3" Connection="Output" /><Destination ToolID="4" Connection="Input" /></Connection>
    <Connection><Origin ToolID="4" Connection="Output" /><Destination ToolID="5" Connection="Input" /></Connection>
  </Connections>
</AlteryxDocument>"""
        wf_file = tmp_path / "transitive_test.yxmd"
        wf_file.write_text(xml, encoding="utf-8")
        wf = parse_workflow(wf_file)
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        adj_mappings = [m for m in sttm.mappings if m.target_attribute == "TotalAdjusted"]
        assert len(adj_mappings) == 2, f"Expected 2 origins (Gross, Discount), got {len(adj_mappings)}"
        src_attrs = {m.source_attribute for m in adj_mappings}
        assert src_attrs == {"Gross", "Discount"}

    def test_banking_domain_synthetic_workflow(self, tmp_path: Path):
        """TEST 15 & Negative Test: Banking/Finance workflow with completely distinct domain terms."""
        xml = """<?xml version="1.0"?>
<AlteryxDocument yxmdVer="2024.1">
  <Nodes>
    <Node ToolID="1"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput" /><Properties><Configuration><File>bank_accounts.csv</File></Configuration></Properties></Node>
    <Node ToolID="2"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput" /><Properties><Configuration><File>credit_ratings.csv</File></Configuration></Properties></Node>
    <Node ToolID="3"><GuiSettings Plugin="AlteryxBasePluginsGui.Join.Join" /><Properties><Configuration>
      <JoinInfo connection="Left"><Field field="AccountNum" /></JoinInfo>
      <JoinInfo connection="Right"><Field field="AccountNum" /></JoinInfo>
    </Configuration></Properties></Node>
    <Node ToolID="4"><GuiSettings Plugin="AlteryxBasePluginsGui.Formula.Formula" /><Properties><Configuration>
      <FormulaFields><FormulaField field="RiskTier" expression="if [CreditScore] &gt; 750 then 'Prime' else 'Subprime' endif" type="V_String" size="20" /></FormulaFields>
    </Configuration></Properties></Node>
    <Node ToolID="5"><GuiSettings Plugin="AlteryxBasePluginsGui.Summarize.Summarize" /><Properties><Configuration>
      <SummarizeFields>
        <SummarizeField field="RiskTier" action="GroupBy" rename="RiskTier" />
        <SummarizeField field="Balance" action="Sum" rename="TotalPortfolioBalance" />
        <SummarizeField field="AccountNum" action="CountDistinct" rename="UniqueAccounts" />
      </SummarizeFields>
    </Configuration></Properties></Node>
    <Node ToolID="6"><GuiSettings Plugin="AlteryxBasePluginsGui.DbFileOutput.DbFileOutput" /><Properties><Configuration><File>portfolio_risk_report.csv</File></Configuration></Properties></Node>
  </Nodes>
  <Connections>
    <Connection><Origin ToolID="1" Connection="Output" /><Destination ToolID="3" Connection="Left" /></Connection>
    <Connection><Origin ToolID="2" Connection="Output" /><Destination ToolID="3" Connection="Right" /></Connection>
    <Connection><Origin ToolID="3" Connection="Join" /><Destination ToolID="4" Connection="Input" /></Connection>
    <Connection><Origin ToolID="4" Connection="Output" /><Destination ToolID="5" Connection="Input" /></Connection>
    <Connection><Origin ToolID="5" Connection="Output" /><Destination ToolID="6" Connection="Input" /></Connection>
  </Connections>
</AlteryxDocument>"""
        wf_file = tmp_path / "banking_test.yxmd"
        wf_file.write_text(xml, encoding="utf-8")
        wf = parse_workflow(wf_file)
        g = build_graph(wf)
        sttm = extract_sttm(wf, g)

        assert sttm.total_mappings == 3
        target_map = {m.target_attribute: m for m in sttm.mappings}
        
        # RiskTier traces to credit_ratings.CreditScore via formula
        assert "RiskTier" in target_map
        assert "credit" in target_map["RiskTier"].source_table.lower()
        assert target_map["RiskTier"].source_attribute == "CreditScore"
        assert target_map["RiskTier"].transformation == "Derived Calculation"

        # UniqueAccounts traces to bank_accounts.AccountNum
        assert "UniqueAccounts" in target_map
        assert "bank" in target_map["UniqueAccounts"].source_table.lower()
        assert target_map["UniqueAccounts"].source_attribute == "AccountNum"
        assert target_map["UniqueAccounts"].transformation == "Aggregation"

    def test_sttm_excel_generation(self, tmp_path: Path):
        """Verify generated Excel workbook formatting, sheets, and headers."""
        wf_path = Path("Demo_Claims_Volume_Extract_reconstructed.yxmd")
        canonical = analyze_canonical(wf_path)
        assert canonical.sttm is not None

        xlsx_path = tmp_path / "Generic_STTM.xlsx"
        generate_sttm_excel(canonical.sttm, xlsx_path)
        assert xlsx_path.exists()

        wb = openpyxl.load_workbook(xlsx_path)
        assert "Source-to-Target Mapping" in wb.sheetnames
        assert "STTM Summary" in wb.sheetnames

        ws1 = wb["Source-to-Target Mapping"]
        headers = [ws1.cell(1, col).value for col in range(1, 7)]
        expected_headers = [
            "Source Table",
            "Source Attribute",
            "Transformation",
            "Transformation Logic",
            "Target Table",
            "Target Attribute",
        ]
        assert headers == expected_headers
        assert ws1.max_row > 10

    def test_sttm_download_endpoint_and_bundle(self):
        """Verify the FastAPI STTM download endpoint and ZIP bundle inclusion."""
        client = TestClient(app)
        wf_path = Path("Demo_Claims_Volume_Extract_reconstructed.yxmd")
        with open(wf_path, "rb") as f:
            resp = client.post("/api/upload", files={"file": ("Demo_Claims.yxmd", f, "application/xml")})

        assert resp.status_code == 200
        analysis_id = resp.json()["analysis_id"]

        sttm_resp = client.get(f"/api/download/{analysis_id}/sttm")
        assert sttm_resp.status_code == 200
        assert "spreadsheetml" in sttm_resp.headers["Content-Type"]

        import zipfile
        zip_resp = client.get(f"/api/download/{analysis_id}/zip")
        assert zip_resp.status_code == 200

        with zipfile.ZipFile(io.BytesIO(zip_resp.content)) as zf:
            filenames = zf.namelist()
            assert any(f.endswith("_STTM.xlsx") or f.endswith("sttm.xlsx") for f in filenames)

    def test_demo_claims_regression(self):
        """TEST 16: Verify Demo Claims produces all expected mappings naturally from generic logic."""
        wf_path = Path("Demo_Claims_Volume_Extract_reconstructed.yxmd")
        canonical = analyze_canonical(wf_path)
        assert canonical.sttm is not None
        sttm = canonical.sttm

        assert sttm.total_mappings == 31

        # Check Claim Count in Product Type, State, and Aging
        claim_count_mappings = [m for m in sttm.mappings if m.target_attribute == "Claim Count"]
        assert len(claim_count_mappings) == 3
        for m in claim_count_mappings:
            assert m.source_table == "Claims Volume"
            assert m.source_attribute == "Claim Number"
            assert m.transformation == "Aggregation"

        # Check Aging Bucket
        aging_m = [m for m in sttm.mappings if m.target_attribute == "Aging Bucket"]
        assert len(aging_m) == 1
        assert aging_m[0].source_table == "Claim Diary Notes"
        assert aging_m[0].source_attribute == "Last Activity Date"

        # Check Total Paid Amount
        paid_m = [m for m in sttm.mappings if m.target_attribute == "Total Paid Amount"]
        assert len(paid_m) == 1
        assert paid_m[0].source_table == "Claim Payments"
        assert paid_m[0].source_attribute == "Payment Amount"

        # Check Product Type
        prod_m = [m for m in sttm.mappings if m.target_attribute == "Product Type"]
        assert len(prod_m) == 1
        assert prod_m[0].source_table == "Policy Master"
        assert prod_m[0].source_attribute == "Product Type"

        # Check State
        state_m = [m for m in sttm.mappings if m.target_attribute == "State"]
        assert len(state_m) == 1
        assert state_m[0].source_table == "Policy Master"
        assert state_m[0].source_attribute == "State"
