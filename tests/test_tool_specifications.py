"""Comprehensive tests for Tool Specifications XLSX document generation, data model, and download endpoints."""

from pathlib import Path
import io
import openpyxl
import pytest
from fastapi.testclient import TestClient

from awa.analysis.workflow_analyzer import analyze_workflow, analyze_canonical
from awa.model.workflow import Workflow, WorkflowMetadata
from awa.model.tool import Tool, Position, ToolConfiguration
from awa.model.connection import Connection
from awa.graph.builder import build_graph, execution_order
from awa.model.tool_specifications import (
    ToolSpecificationRow,
    ToolSpecificationsDocument,
    build_tool_specifications_document,
    format_input_tools,
    format_output_tools,
)
from awa.generators.tool_specifications_generator import generate_tool_specifications_excel
from awa.llm.client import FakeLLMClient, set_default_llm_client
from awa.llm.generator import LLMNarrativeGenerator, set_default_generator
from awa.llm.cache import LLMNarrativeCache
from backend.app.main import app


class TestToolSpecifications:
    """Validate Tool Specifications XLSX generation against the architectural contract."""

    EXPECTED_HEADERS = [
        "Tool ID",
        "XML Tool Name",
        "Tool Type",
        "Role — What It Does",
        "Data Flow Explanation",
        "Input Tool",
        "Output Tool",
    ]

    def test_tool_specifications_excel_structure(self, tmp_path: Path):
        """Verify XLSX creation, worksheet title, column headers, ordering, and formatting."""
        canonical = analyze_canonical(Path("fixtures/basic/simple_filter.yxmd"))
        wf = canonical.workflow
        graph = canonical.graph

        tool_doc = build_tool_specifications_document(wf, graph)
        out_file = tmp_path / "Tool Specifications.xlsx"
        generate_tool_specifications_excel(tool_doc, out_file)

        assert out_file.exists()

        wb = openpyxl.load_workbook(out_file)
        assert "Tool Specifications" in wb.sheetnames
        ws = wb["Tool Specifications"]

        # 1. Header row verification
        actual_headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        assert actual_headers == self.EXPECTED_HEADERS

        # 2. Frozen header pane
        assert ws.freeze_panes == "A2"

        # 3. Row count matches tool count exactly
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == len(wf.tools)

        # 4. Check column values for each tool
        for row in data_rows:
            tool_id_str, xml_name, tool_type, role, data_flow, in_tool, out_tool = row
            assert tool_id_str.startswith("#")
            assert xml_name != ""
            assert tool_type != ""
            assert role != ""
            assert data_flow != ""
            assert in_tool != ""
            assert out_tool != ""

    def test_tool_specifications_deterministic_topology_facts(self, tmp_path: Path):
        """Verify that Input Tool and Output Tool correctly reflect graph boundaries and multi-connections."""
        wf = Workflow(
            metadata=WorkflowMetadata(name="Branching Workflow", version="2024.1"),
            tools={
                1: Tool(
                    tool_id=1,
                    plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput",
                    tool_type="DbFileInput",
                    name="Input Orders",
                    position=Position(0, 0),
                    configuration=ToolConfiguration(raw_xml="<Configuration/>", parsed={"file_path": "orders.csv"}),
                ),
                2: Tool(
                    tool_id=2,
                    plugin="AlteryxBasePluginsGui.DbFileInput.DbFileInput",
                    tool_type="DbFileInput",
                    name="Input Customers",
                    position=Position(0, 100),
                    configuration=ToolConfiguration(raw_xml="<Configuration/>", parsed={"file_path": "customers.csv"}),
                ),
                3: Tool(
                    tool_id=3,
                    plugin="AlteryxBasePluginsGui.Join.Join",
                    tool_type="Join",
                    name="Join Orders Customers",
                    position=Position(100, 50),
                    configuration=ToolConfiguration(raw_xml="<Configuration/>", parsed={"join_fields": ["CustomerID"]}),
                ),
                4: Tool(
                    tool_id=4,
                    plugin="AlteryxBasePluginsGui.DbFileOutput.DbFileOutput",
                    tool_type="DbFileOutput",
                    name="Output Matched",
                    position=Position(200, 0),
                    configuration=ToolConfiguration(raw_xml="<Configuration/>", parsed={"file_path": "matched.xlsx"}),
                ),
                5: Tool(
                    tool_id=5,
                    plugin="AlteryxBasePluginsGui.DbFileOutput.DbFileOutput",
                    tool_type="DbFileOutput",
                    name="Output Unmatched",
                    position=Position(200, 100),
                    configuration=ToolConfiguration(raw_xml="<Configuration/>", parsed={"file_path": "unmatched.xlsx"}),
                ),
            },
            connections=[
                Connection(origin_tool_id=1, origin_anchor="Output", destination_tool_id=3, destination_anchor="Left"),
                Connection(origin_tool_id=2, origin_anchor="Output", destination_tool_id=3, destination_anchor="Right"),
                Connection(origin_tool_id=3, origin_anchor="Join", destination_tool_id=4, destination_anchor="Input"),
                Connection(origin_tool_id=3, origin_anchor="LeftUnassigned", destination_tool_id=5, destination_anchor="Input"),
            ],
        )
        graph = build_graph(wf)

        doc = build_tool_specifications_document(wf, graph)
        row_map = {r.tool_id: r for r in doc.rows}

        # Tool 1: Source tool (No upstream)
        assert row_map[1].input_tool == "Source"
        assert row_map[1].output_tool == "#3 Join"

        # Tool 2: Source tool (No upstream)
        assert row_map[2].input_tool == "Source"
        assert row_map[2].output_tool == "#3 Join"

        # Tool 3: Join (Multiple inputs: Tool 1 and Tool 2, Multiple outputs: Tool 4 and Tool 5)
        assert "#1 DbFileInput" in row_map[3].input_tool
        assert "#2 DbFileInput" in row_map[3].input_tool
        assert "#4 DbFileOutput" in row_map[3].output_tool
        assert "#5 DbFileOutput" in row_map[3].output_tool

        # Tool 4 & 5: Terminal tools (No downstream)
        assert row_map[4].input_tool == "#3 Join"
        assert row_map[4].output_tool == "None"
        assert row_map[5].input_tool == "#3 Join"
        assert row_map[5].output_tool == "None"

    def test_llm_generated_role_and_data_flow_integration(self, tmp_path: Path):
        """Verify LLM-generated role and data-flow explanations populate into the XLSX document."""
        import json

        def mock_llm(system: str, user: str) -> str | None:
            if "Filter" in user:
                return json.dumps({
                    "tool_id": 2,
                    "role": "Filters incoming customer transactions to isolate active accounts for downstream processing.",
                    "data_flow_explanation": "Receives raw customer records from #1 DbFileInput, applies the active status predicate, and routes qualifying records to #3 DbFileOutput.",
                })
            elif "DbFileInput" in user:
                return json.dumps({
                    "tool_id": 1,
                    "role": "Ingests daily customer transaction records from the source CSV file into the processing pipeline.",
                    "data_flow_explanation": "Reads records from customers.csv and passes the initial uncompressed dataset to #2 Filter.",
                })
            elif "DbFileOutput" in user:
                return json.dumps({
                    "tool_id": 3,
                    "role": "Publishes the filtered active customer dataset to the finalized Excel reporting deliverable.",
                    "data_flow_explanation": "Receives qualifying active records from #2 Filter and writes the finalized matrix to active_customers.xlsx.",
                })
            return None

        client = FakeLLMClient(generator_fn=mock_llm)
        gen = LLMNarrativeGenerator(client=client, cache=LLMNarrativeCache())
        set_default_llm_client(client)
        set_default_generator(gen)

        try:
            canonical = analyze_canonical(Path("fixtures/basic/simple_filter.yxmd"))
            wf = canonical.workflow
            graph = canonical.graph

            tool_specs = gen.generate_all_tool_specifications(wf, graph=graph, workflow_id="test_filter")
            assert len(tool_specs) == len(wf.tools)

            doc = build_tool_specifications_document(wf, graph, tool_specs=tool_specs)
            out_file = tmp_path / "Tool Specifications.xlsx"
            generate_tool_specifications_excel(doc, out_file)

            wb = openpyxl.load_workbook(out_file)
            ws = wb["Tool Specifications"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            filter_row = next((r for r in rows if r[2] == "Filter"), None)
            assert filter_row is not None
            role = filter_row[3]
            data_flow = filter_row[4]

            assert "Filters incoming customer transactions" in role
            assert "Receives raw customer records from #1 DbFileInput" in data_flow
        finally:
            set_default_llm_client(None)
            set_default_generator(None)

    def test_multi_domain_regression_and_no_cross_domain_leakage(self, tmp_path: Path):
        """Verify that a generic filter workflow produces no claims-demo terminology."""
        canonical = analyze_canonical(Path("fixtures/basic/simple_filter.yxmd"))
        tool_doc = build_tool_specifications_document(canonical.workflow, canonical.graph)
        out_file = tmp_path / "Tool_Specifications.xlsx"
        generate_tool_specifications_excel(tool_doc, out_file)

        wb = openpyxl.load_workbook(out_file)
        ws = wb["Tool Specifications"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        all_text = " ".join(" ".join(str(c) for c in r) for r in rows)

        forbidden_claims_terms = [
            "Claims_Volume_Extract_Demo",
            "Policy_Master_Demo",
            "Claim_Payments_Demo",
            "policyholder",
            "actuarial",
        ]
        for term in forbidden_claims_terms:
            assert term not in all_text, f"Found forbidden claims term '{term}' in generic filter tool specs!"

    def test_download_endpoint_and_tools_page_url(self):
        """Verify download endpoint returns the Tool Specifications Excel file."""
        client = TestClient(app)
        wf_path = Path("Demo_Claims_Volume_Extract_reconstructed.yxmd")
        with open(wf_path, "rb") as f:
            resp = client.post("/api/upload", files={"file": ("Demo_Claims.yxmd", f, "application/xml")})

        assert resp.status_code == 200
        analysis_id = resp.json()["analysis_id"]

        # Download Tool Specifications XLSX
        resp_tool = client.get(f"/api/download/{analysis_id}/tool-specifications")
        assert resp_tool.status_code == 200
        assert "Tool_Specifications.xlsx" in resp_tool.headers.get("Content-Disposition", "")
        wb = openpyxl.load_workbook(io.BytesIO(resp_tool.content))
        assert "Tool Specifications" in wb.sheetnames
        ws = wb["Tool Specifications"]
        assert ws.cell(row=1, column=1).value == "Tool ID"
        assert ws.cell(row=1, column=4).value == "Role — What It Does"
        assert ws.cell(row=1, column=5).value == "Data Flow Explanation"
        assert ws.cell(row=1, column=6).value == "Input Tool"
        assert ws.cell(row=1, column=7).value == "Output Tool"
